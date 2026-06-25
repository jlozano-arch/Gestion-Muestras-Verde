import csv
import json
import logging
import os
import re
import unicodedata
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import httpx
from openpyxl import load_workbook


ERP_DATA_PATH_ENV = "ERP_DATA_PATH"
ERP_SOURCE_ENV = "ERP_SOURCE"
ERP_APPS_SCRIPT_URL_ENV = "ERP_APPS_SCRIPT_URL"
ERP_SOURCE_FILE = "file"
ERP_SOURCE_APPS_SCRIPT = "apps_script"
ERP_LOG = logging.getLogger("app.erp_integration")
ERP_BASIC_TIMEOUT_SECONDS = 8.0
ERP_TRACE_TIMEOUT_SECONDS = 25.0

FIELD_ALIASES = {
    "cvc": {"cvc", "ctr compra", "ctr. compra", "contrato compra", "contrato de compra", "purchase contract"},
    "provider": {"proveedor", "supplier", "provider"},
    "supplier_reference": {"ref proveedor", "ref. proveedor", "referencia proveedor", "supplier reference", "supplier ref"},
    "quality": {"calidad", "quality", "description", "descripcion"},
    "origin": {"pais", "country", "origen", "origin", "pais origen"},
    "warehouse": {"almacen", "warehouse"},
    "warehouse_lot": {"almacen long no lote", "almacen long numero lote", "almacen lote", "no lote", "numero lote"},
    "sample_reference": {"muestra", "sample"},
    "bags_purchased": {"cantidad sacos", "sacos comprados", "bags purchased", "purchased bags", "sacos contrato"},
    "kg_purchased": {"kg teorico", "kg comprados", "kilos comprados", "purchased kg", "kg purchased", "contract kg"},
    "bags_available": {"stock sacos", "sacos disponibles", "available bags", "bags available"},
    "kg_available": {"kg disponibles", "kilos disponibles", "available kg", "kg available"},
    "purchase_price": {"precio fijo", "precio compra", "purchase price", "price", "precio"},
    "contract_date": {"fecha contrato", "contract date", "fecha compra", "purchase date"},
    "contract_status": {"estado", "estado contrato", "contract status", "status"},
    "incoterm": {"incoterm"},
    "comments": {"comentarios", "comments", "observaciones", "notes"},
}

RAW_APPS_SCRIPT_FIELD_MAP = {
    "cvc": "cvc",
    "proveedor": "provider",
    "ref_proveedor": "supplier_reference",
    "calidad": "quality",
    "pais_origen": "origin",
    "almacen": "warehouse",
    "almacen_lote": "warehouse_lot",
    "muestra": "sample_reference",
    "sacos_comprados": "bags_purchased",
    "kg_comprados": "kg_purchased",
    "sacos_disponibles": "bags_available",
    "kg_disponibles": "kg_available",
    "stock_sacos": "stock_bags",
    "precio_compra": "purchase_price",
    "fecha_contrato": "contract_date",
    "estado_contrato": "contract_status",
    "estado": "contract_status",
    "incoterm": "incoterm",
    "comentarios": "comments",
}

RAW_APPS_SCRIPT_SALE_FIELD_MAP = {
    "cvv": "cvv",
    "cliente": "client",
    "calidad_venta": "sale_quality",
    "sacos_vendidos": "bags_sold",
    "kg_vendidos": "kg_sold",
    "precio_venta": "sale_price",
    "valor_contrato": "contract_value",
    "fecha_factura": "invoice_date",
    "cobro": "payment_status",
    "incoterm_venta": "sale_incoterm",
    "comentarios_venta": "sale_comments",
}

RAW_APPS_SCRIPT_SUMMARY_FIELD_MAP = {
    "total_kg_vendidos": "total_kg_sold",
    "total_sacos_vendidos": "total_bags_sold",
    "numero_cvv": "cvv_count",
    "clientes_unicos": "unique_clients",
    "kg_comprados": "kg_purchased",
    "kg_pendientes": "kg_pending",
    "precio_compra": "purchase_price",
    "precio_venta_medio": "average_sale_price",
    "margen_bruto_medio": "average_gross_margin",
}

RAW_APPS_SCRIPT_MOVEMENT_FIELD_MAP = {
    "origen": "source",
    "fecha": "date",
    "tipo_operacion": "operation_type",
    "cvc": "cvc",
    "cvv": "cvv",
    "cliente_proveedor": "counterparty",
    "sacos": "bags",
    "calidad": "quality",
    "estado": "status",
}

RAW_APPS_SCRIPT_TRACEABILITY_SUMMARY_FIELD_MAP = {
    "total_sacos_movidos": "total_bags_moved",
    "sacos_entrada": "bags_in",
    "sacos_salida": "bags_out",
    "sacos_app": "bags_app",
    "sacos_deapp": "bags_deapp",
    "clientes_unicos": "unique_clients",
    "numero_cvv": "cvv_count",
}

DISPLAY_LABELS = {
    "cvc": "CVC",
    "provider": "Proveedor ERP",
    "supplier_reference": "Ref. proveedor ERP",
    "quality": "Calidad ERP",
    "origin": "Pais/origen ERP",
    "warehouse": "Almacen ERP",
    "warehouse_lot": "Almacen / lote ERP",
    "sample_reference": "Muestra ERP",
    "bags_purchased": "Sacos comprados",
    "kg_purchased": "Kg comprados",
    "bags_available": "Sacos disponibles",
    "stock_bags": "Stock sacos",
    "kg_available": "Kg disponibles",
    "purchase_price": "Precio compra",
    "contract_date": "Fecha contrato",
    "contract_status": "Estado contrato",
    "incoterm": "Incoterm",
    "comments": "Comentarios ERP",
}

SALE_LABELS = {
    "cvv": "CVV",
    "client": "Cliente",
    "sale_quality": "Calidad venta",
    "bags_sold": "Sacos vendidos",
    "kg_sold": "Kg vendidos",
    "sale_price": "Precio venta",
    "contract_value": "Valor contrato",
    "invoice_date": "Fecha factura",
    "payment_status": "Cobro",
    "sale_incoterm": "Incoterm venta",
    "sale_comments": "Comentarios venta",
}

SUMMARY_LABELS = {
    "total_kg_sold": "Kg vendidos",
    "total_bags_sold": "Sacos vendidos",
    "cvv_count": "Numero CVV",
    "unique_clients": "Clientes unicos",
    "kg_purchased": "Kg comprados",
    "kg_pending": "Kg pendientes",
    "purchase_price": "Precio compra",
    "average_sale_price": "Precio venta medio",
    "average_gross_margin": "Margen bruto medio",
    "total_bags_moved": "Sacos movidos",
    "bags_in": "Sacos entrada",
    "bags_out": "Sacos salida",
    "bags_app": "Sacos APP",
    "bags_deapp": "Sacos DEAPP",
}

MOVEMENT_LABELS = {
    "source": "Origen",
    "date": "Fecha",
    "operation_type": "Operacion",
    "cvc": "CVC",
    "cvv": "CVV",
    "counterparty": "Cliente/proveedor",
    "bags": "Sacos",
    "quality": "Calidad",
    "status": "Estado",
}

PRIVATE_SALE_FIELDS = [
    "cvv",
    "client",
    "sale_quality",
    "bags_sold",
    "kg_sold",
    "sale_price",
    "contract_value",
    "invoice_date",
    "payment_status",
    "sale_incoterm",
    "sale_comments",
]

PUBLIC_SUMMARY_FIELDS = [
    "cvv_count",
    "total_kg_sold",
    "kg_pending",
]

PRIVATE_SUMMARY_FIELDS = [
    "cvv_count",
    "unique_clients",
    "kg_purchased",
    "total_kg_sold",
    "kg_pending",
    "total_bags_sold",
    "purchase_price",
    "average_sale_price",
    "average_gross_margin",
]

PRIVATE_TRACEABILITY_SUMMARY_FIELDS = [
    "cvv_count",
    "total_bags_moved",
    "bags_in",
    "bags_out",
    "bags_app",
    "bags_deapp",
    "unique_clients",
]

PUBLIC_TRACEABILITY_SUMMARY_FIELDS = [
    "cvv_count",
    "total_bags_moved",
    "bags_out",
]

PRIVATE_MOVEMENT_FIELDS = [
    "source",
    "date",
    "operation_type",
    "cvc",
    "cvv",
    "counterparty",
    "bags",
    "quality",
    "status",
]


def normalize_cvc(cvc: str | None) -> str:
    if cvc is None:
        return ""
    text = str(cvc).strip().upper()
    text = re.sub(r"\s+", " ", text)
    return "" if text in {"", "-", "N/A", "NA", "NONE", "NULL", "NAN"} else text


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"[._:/\\()\[\]-]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip().lower()


NORMALIZED_ALIASES = {
    field: {_normalize_text(alias) for alias in aliases}
    for field, aliases in FIELD_ALIASES.items()
}

APPS_SCRIPT_FIELD_MAP = {
    _normalize_text(raw_key): field
    for raw_key, field in RAW_APPS_SCRIPT_FIELD_MAP.items()
}

APPS_SCRIPT_SALE_FIELD_MAP = {
    _normalize_text(raw_key): field
    for raw_key, field in RAW_APPS_SCRIPT_SALE_FIELD_MAP.items()
}

APPS_SCRIPT_SUMMARY_FIELD_MAP = {
    _normalize_text(raw_key): field
    for raw_key, field in RAW_APPS_SCRIPT_SUMMARY_FIELD_MAP.items()
}

APPS_SCRIPT_MOVEMENT_FIELD_MAP = {
    _normalize_text(raw_key): field
    for raw_key, field in RAW_APPS_SCRIPT_MOVEMENT_FIELD_MAP.items()
}

APPS_SCRIPT_TRACEABILITY_SUMMARY_FIELD_MAP = {
    _normalize_text(raw_key): field
    for raw_key, field in RAW_APPS_SCRIPT_TRACEABILITY_SUMMARY_FIELD_MAP.items()
}


def _field_for_header(header: Any) -> str | None:
    normalized = _normalize_text(header)
    if not normalized:
        return None
    for field, aliases in NORMALIZED_ALIASES.items():
        if normalized in aliases:
            return field
    return None


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return "" if text.lower() in {"", "-", "none", "null", "nan", "n/a", "na"} else text


def _mapped_record(headers: list[Any], values: list[Any], source_sheet: str | None = None) -> dict[str, str]:
    record: dict[str, str] = {}
    for index, header in enumerate(headers):
        field = _field_for_header(header)
        if not field or field in record:
            continue
        value = values[index] if index < len(values) else None
        cleaned = _clean_value(value)
        if cleaned:
            record[field] = normalize_cvc(cleaned) if field == "cvc" else cleaned
    if source_sheet:
        record["source_sheet"] = source_sheet
    return record


def _read_csv(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        reader = csv.reader(handle, dialect)
        all_rows = list(reader)
    if not all_rows:
        return rows
    headers = all_rows[0]
    for values in all_rows[1:]:
        record = _mapped_record(headers, values)
        if record:
            rows.append(record)
    return rows


def _find_header_row(sheet) -> tuple[int, list[Any]] | None:
    for row_index, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, values_only=True), start=1):
        mapped_fields = {_field_for_header(value) for value in row}
        mapped_fields.discard(None)
        if "cvc" in mapped_fields and len(mapped_fields) >= 2:
            return row_index, list(row)
    return None


def _read_excel(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        header = _find_header_row(sheet)
        if not header:
            continue
        header_row_index, headers = header
        for values in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
            record = _mapped_record(headers, list(values), source_sheet=sheet_name)
            if record:
                rows.append(record)
    workbook.close()
    return rows


def _read_erp_records(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt", ".tsv"}:
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_excel(path)
    return []


def _status(status: str, **extra: Any) -> dict[str, Any]:
    return {"status": status, **extra}


def _normalize_record_with_map(record: dict[str, Any] | None, field_map: dict[str, str]) -> dict[str, str]:
    normalized = {}
    for raw_key, raw_value in (record or {}).items():
        key = field_map.get(_normalize_text(raw_key))
        if not key:
            continue
        value = _clean_value(raw_value)
        if value:
            normalized[key] = normalize_cvc(value) if key == "cvc" else value
    return normalized


def _normalize_apps_script_record(record: dict[str, Any] | None) -> dict[str, str]:
    return _normalize_record_with_map(record, APPS_SCRIPT_FIELD_MAP)


def _normalize_apps_script_sale(record: dict[str, Any] | None) -> dict[str, str]:
    return _normalize_record_with_map(record, APPS_SCRIPT_SALE_FIELD_MAP)


def _normalize_apps_script_summary(record: dict[str, Any] | None) -> dict[str, str]:
    return _normalize_record_with_map(record, APPS_SCRIPT_SUMMARY_FIELD_MAP)


def _normalize_apps_script_movement(record: dict[str, Any] | None) -> dict[str, str]:
    return _normalize_record_with_map(record, APPS_SCRIPT_MOVEMENT_FIELD_MAP)


def _normalize_apps_script_traceability_summary(record: dict[str, Any] | None) -> dict[str, str]:
    return _normalize_record_with_map(record, APPS_SCRIPT_TRACEABILITY_SUMMARY_FIELD_MAP)


def _log_apps_script_response(url: str, response: httpx.Response, payload: Any = None) -> None:
    ERP_LOG.warning("ERP Apps Script URL consultada: %s", url)
    ERP_LOG.warning("ERP Apps Script status_code: %s", response.status_code)
    ERP_LOG.warning("ERP Apps Script content-type: %s", response.headers.get("content-type", ""))
    ERP_LOG.warning("ERP Apps Script response.text[0:500]: %s", response.text[:500])
    ERP_LOG.warning("ERP Apps Script response.json(): %s", payload)


def _apps_script_request(cvc: str, action: str, timeout_seconds: float) -> dict | None:
    normalized_cvc = normalize_cvc(cvc)
    endpoint = os.getenv(ERP_APPS_SCRIPT_URL_ENV, "").strip()
    if not endpoint:
        return _status("not_configured", message=f"{ERP_APPS_SCRIPT_URL_ENV} no configurado")
    if not normalized_cvc:
        return _status("no_cvc")

    separator = "&" if "?" in endpoint else "?"
    url = f"{endpoint}{separator}{urlencode({'action': action, 'cvc': normalized_cvc})}"
    try:
        with httpx.Client(follow_redirects=True, timeout=timeout_seconds, headers={"Accept": "application/json"}) as client:
            response = client.get(url)
        try:
            payload = response.json()
        except Exception:
            payload = json.loads(response.text)
        _log_apps_script_response(url, response, payload)
        response.raise_for_status()
    except Exception as exc:
        return _status("error", message=f"No se pudo consultar Apps Script: {exc}", source="google_sheets_largos")

    status = _clean_value(payload.get("status")).lower()
    source = _clean_value(payload.get("source")) or "google_sheets_largos"
    if status == "not_found":
        return _status(
            "not_found",
            cvc=normalize_cvc(payload.get("cvc") or normalized_cvc),
            source=source,
            warnings=payload.get("warnings") or [],
        )
    if status == "found":
        return _status(
            "found",
            cvc=normalize_cvc(payload.get("cvc") or normalized_cvc),
            data=_normalize_apps_script_record(payload.get("data")),
            sales=[_normalize_apps_script_sale(item) for item in payload.get("ventas_asociadas") or []],
            commercial_summary=_normalize_apps_script_summary(payload.get("resumen_comercial")),
            traceability_movements=[
                _normalize_apps_script_movement(item)
                for item in payload.get("trazabilidad_movimientos") or []
            ],
            associated_cvvs=[_clean_value(item) for item in payload.get("cvv_asociados") or [] if _clean_value(item)],
            traceability_summary=_normalize_apps_script_traceability_summary(payload.get("resumen_trazabilidad")),
            source=source,
            warnings=payload.get("warnings") or [],
        )
    if status == "multiple":
        matches = [_normalize_apps_script_record(item) for item in payload.get("matches") or []]
        return _status(
            "multiple",
            cvc=normalize_cvc(payload.get("cvc") or normalized_cvc),
            count=len(matches),
            matches=matches,
            source=source,
            warnings=payload.get("warnings") or [],
        )
    if status == "error":
        return _status(
            "error",
            message=_clean_value(payload.get("message")) or "Error devuelto por Apps Script",
            source=source,
            warnings=payload.get("warnings") or [],
        )
    return _status("error", message=f"Respuesta ERP no reconocida: {status or 'sin status'}", source=source)


def _apps_script_lookup(cvc: str) -> dict | None:
    return _apps_script_request(cvc, "getDatosMuestraERP", ERP_BASIC_TIMEOUT_SECONDS)


def _apps_script_trace_lookup(cvc: str) -> dict | None:
    result = _apps_script_request(cvc, "getTrazabilidadMuestraERP", ERP_TRACE_TIMEOUT_SECONDS)
    if result and result.get("status") == "error" and "timed out" in (result.get("message") or "").lower():
        result["message"] = "Trazabilidad no disponible temporalmente"
    return result


def _file_lookup(cvc: str) -> dict | None:
    normalized_cvc = normalize_cvc(cvc)
    erp_path = os.getenv(ERP_DATA_PATH_ENV, "").strip()
    if not erp_path:
        return _status("not_configured")
    path = Path(erp_path)
    if not path.exists() or not path.is_file():
        return _status("not_configured", message=f"{ERP_DATA_PATH_ENV} no apunta a un archivo valido")
    if not normalized_cvc:
        return _status("no_cvc")

    try:
        records = _read_erp_records(path)
    except Exception as exc:
        return _status("not_configured", message=f"No se pudo leer {ERP_DATA_PATH_ENV}: {exc}")
    matches = [record for record in records if normalize_cvc(record.get("cvc")) == normalized_cvc]
    if not matches:
        return _status("not_found", cvc=normalized_cvc, source=ERP_SOURCE_FILE)
    if len(matches) > 1:
        return _status("multiple", cvc=normalized_cvc, count=len(matches), matches=matches, source=ERP_SOURCE_FILE)
    return _status("found", cvc=normalized_cvc, data=matches[0], source=ERP_SOURCE_FILE)


def get_erp_data_by_cvc(cvc: str) -> dict | None:
    """Read-only lookup of ERP data by exact normalized CVC."""
    source = os.getenv(ERP_SOURCE_ENV, ERP_SOURCE_FILE).strip().lower()
    if source == ERP_SOURCE_APPS_SCRIPT:
        return _apps_script_lookup(cvc)
    return _file_lookup(cvc)


def get_erp_trace_by_cvc(cvc: str) -> dict | None:
    """Read-only lookup of slower ERP traceability data by normalized CVC."""
    source = os.getenv(ERP_SOURCE_ENV, ERP_SOURCE_FILE).strip().lower()
    if source == ERP_SOURCE_APPS_SCRIPT:
        return _apps_script_trace_lookup(cvc)
    return _status("not_configured", message="La trazabilidad ERP requiere ERP_SOURCE=apps_script")


def erp_display_rows(data: dict | None, public: bool = False) -> list[dict[str, str]]:
    if not data or data.get("status") != "found":
        return []
    record = data.get("data") or {}
    fields = [
        "quality",
        "supplier_reference",
        "warehouse_lot",
        "sample_reference",
        "bags_available",
        "stock_bags",
        "kg_available",
        "contract_status",
    ] if public else [
        "cvc",
        "provider",
        "supplier_reference",
        "quality",
        "warehouse_lot",
        "sample_reference",
        "bags_purchased",
        "kg_purchased",
        "bags_available",
        "stock_bags",
        "kg_available",
        "purchase_price",
        "incoterm",
        "contract_status",
        "comments",
    ]
    rows = []
    for field in fields:
        value = _clean_value(record.get(field))
        if value:
            rows.append({"label": DISPLAY_LABELS[field], "value": value})
    return rows


def erp_commercial_summary_rows(data: dict | None, public: bool = False) -> list[dict[str, str]]:
    if not data or data.get("status") != "found":
        return []
    summary = data.get("commercial_summary") or {}
    fields = PUBLIC_SUMMARY_FIELDS if public else PRIVATE_SUMMARY_FIELDS
    rows = []
    for field in fields:
        value = _clean_value(summary.get(field))
        if value:
            rows.append({"label": SUMMARY_LABELS[field], "value": value})
    if public:
        status_value = _clean_value((data.get("data") or {}).get("contract_status"))
        if status_value:
            rows.append({"label": DISPLAY_LABELS["contract_status"], "value": status_value})
    return rows


def erp_sales_rows(data: dict | None) -> list[dict[str, str]]:
    if not data or data.get("status") != "found":
        return []
    rows = []
    for sale in data.get("sales") or []:
        row = {}
        for field in PRIVATE_SALE_FIELDS:
            value = _clean_value(sale.get(field))
            if value:
                row[field] = value
        if row:
            rows.append(row)
    return rows


def erp_sales_columns() -> list[dict[str, str]]:
    return [{"field": field, "label": SALE_LABELS[field]} for field in PRIVATE_SALE_FIELDS]


def erp_traceability_summary_rows(data: dict | None, public: bool = False) -> list[dict[str, str]]:
    if not data or data.get("status") != "found":
        return []
    summary = data.get("traceability_summary") or {}
    fields = PUBLIC_TRACEABILITY_SUMMARY_FIELDS if public else PRIVATE_TRACEABILITY_SUMMARY_FIELDS
    rows = []
    for field in fields:
        value = _clean_value(summary.get(field))
        if value:
            rows.append({"label": SUMMARY_LABELS[field], "value": value})
    return rows


def erp_traceability_movement_rows(data: dict | None) -> list[dict[str, str]]:
    if not data or data.get("status") != "found":
        return []
    rows = []
    for movement in data.get("traceability_movements") or []:
        row = {}
        for field in PRIVATE_MOVEMENT_FIELDS:
            value = _clean_value(movement.get(field))
            if value:
                row[field] = value
        if row:
            rows.append(row)
    return rows


def erp_traceability_movement_columns() -> list[dict[str, str]]:
    return [{"field": field, "label": MOVEMENT_LABELS[field]} for field in PRIVATE_MOVEMENT_FIELDS]
