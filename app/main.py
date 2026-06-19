from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.requests import Request
from sqlalchemy.orm import Session
from sqlalchemy import desc, func, or_
from datetime import datetime
from pathlib import Path
import qrcode
import io
from reportlab.lib.pagesizes import letter, A6, A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib.colors import HexColor
import json
import os
import shutil
import hashlib
import re
import unicodedata
from urllib.parse import quote_plus

from .database import get_db, create_tables
from .models import Sample, Tasting, Shipment, Event, Document, SampleStatus, ImportBatch, ImportRow
from .countries import get_country_name, get_country_flag, get_all_countries

# Create tables on startup
create_tables()

# Create FastAPI app
app = FastAPI(
    title="Gestión de Muestras de Café Verde",
    description="Indian Ecotrade - Sistema de gestión de muestras de café verde",
    version="1.0.0"
)

# Mount static files
static_path = Path(__file__).parent / "static"
static_path.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_path)), name="static")

# Configurable uploads path
UPLOADS_DIR = os.getenv("UPLOADS_DIR", "uploads")
uploads_path = Path(UPLOADS_DIR)
uploads_path.mkdir(parents=True, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(uploads_path)), name="uploads")

# Setup templates
templates_path = Path(__file__).parent / "templates"
templates = Jinja2Templates(directory=str(templates_path))

STATUS_LABELS = {
    "received": "Recibida",
    "available": "Disponible",
    "approved": "Aprobada",
    "rejected": "Rechazada",
    "shipped": "Enviada",
    "archived": "Archivada",
}

LEGACY_STATUS_ALIASES = {
    "pending": "received",
    "pendiente": "received",
    "analyzing": "received",
    "evaluated": "approved",
    "purchased": "approved",
    "partially_shipped": "available",
    "partially-shipped": "available",
    "exhausted": "shipped",
    "agotada": "shipped",
}

TASTING_RESULT_LABELS = {
    "pending": "Pendiente",
    "approved": "Aprobada",
    "rejected": "Rechazada",
}


def normalize_status_value(status) -> str:
    if status is None:
        return ""
    if hasattr(status, 'value'):
        status = status.value
    status = str(status).strip().lower()
    return LEGACY_STATUS_ALIASES.get(status, status)


def status_label(status) -> str:
    status = normalize_status_value(status)
    return STATUS_LABELS.get(status, status.replace('_', ' ').title())


def status_class(status) -> str:
    status = normalize_status_value(status)
    if not status:
        return "unknown"
    return status.lower().replace('_', '-').replace(' ', '-')


def tasting_result_label(result) -> str:
    if result is None:
        return "Pendiente"
    if hasattr(result, 'value'):
        result = result.value
    return TASTING_RESULT_LABELS.get(str(result), str(result).title())


def _sample_country_display(sample: Sample) -> str:
    country_name = display_value(sample.country_name, "")
    if country_name:
        return country_name
    if sample.country_code and get_country_name(sample.country_code):
        return get_country_name(sample.country_code)
    origin_country = _origin_country_candidate(sample.origin)
    if origin_country:
        return origin_country
    return ""


def display_value(value, default="-"):
    if value is None:
        return default
    text = str(value).strip()
    if not text or text.lower() in {"none", "null", "nan", "-"}:
        return default
    return text


def _data_option_label(value) -> str:
    label = display_value(value, "")
    if not label:
        return ""
    return re.sub(r"\s+", " ", label).strip()


def _sample_field_options(db: Session, field_name: str) -> list[dict]:
    seen = {}
    for (value,) in db.query(getattr(Sample, field_name)).distinct().all():
        label = _data_option_label(value)
        if not label:
            continue
        key = _normalize_origin_key(label)
        if key and key not in seen:
            seen[key] = label
    return [
        {"value": label, "label": label}
        for label in sorted(seen.values(), key=lambda value: _normalize_origin_key(value))
    ]


def _origin_country_candidate(origin) -> str:
    key = _normalize_origin_key(origin)
    if not key:
        return ""
    aliases = [
        ("costa de marfil", "Costa de Marfil"),
        ("cote d ivoire", "Costa de Marfil"),
        ("cote divoire", "Costa de Marfil"),
        ("ivory coast", "Costa de Marfil"),
        ("sierra leona", "Sierra Leona"),
        ("sierra leone", "Sierra Leona"),
        ("brasil", "Brasil"),
        ("brazil", "Brasil"),
        ("vietnam", "Vietnam"),
        ("viet nam", "Vietnam"),
        ("vietnem", "Vietnam"),
        ("uganda", "Uganda"),
        ("colombia", "Colombia"),
        ("peru", "Perú"),
        ("mexico", "Mexico"),
        ("honduras", "Honduras"),
        ("india", "India"),
        ("ruanda", "Ruanda"),
        ("rwanda", "Ruanda"),
        ("tanzania", "Tanzania"),
        ("angola", "Angola"),
        ("cameroun", "Cameroun"),
        ("cameroon", "Cameroun"),
        ("congo", "Congo"),
        ("cuba", "Cuba"),
        ("guinea", "Guinea"),
        ("venezuela", "Venezuela"),
    ]
    for alias, label in aliases:
        if key == alias or key.startswith(f"{alias} "):
            return label
    return ""


def _sample_country_options(db: Session) -> list[dict]:
    seen = {}
    for sample in db.query(Sample).all():
        display = _sample_country_display(sample)
        if not display:
            continue
        key = _normalize_origin_key(display)
        if key and key not in seen:
            seen[key] = display
    return [
        {"value": display, "label": display}
        for display in sorted(seen.values(), key=lambda value: _normalize_origin_key(value))
    ]

logo_exists = (Path(__file__).parent / "static" / "logo.png").exists()

templates.env.globals["status_label"] = status_label
templates.env.globals["status_class"] = status_class
templates.env.globals["tasting_result_label"] = tasting_result_label
templates.env.globals["sample_country_display"] = _sample_country_display
templates.env.globals["display_value"] = display_value
templates.env.globals["logo_exists"] = logo_exists


def _safe_delete_upload_file(file_path):
    if not file_path:
        return False
    try:
        uploads_root = Path(UPLOADS_DIR).resolve()
        target = Path(file_path).resolve()
        if not target.is_file():
            return False
        if uploads_root == target or uploads_root not in target.parents:
            return False
        target.unlink()
        return True
    except Exception:
        return False


def _delete_sample_records(db: Session, sample: Sample, import_final_action: str = None, import_batch_id: int = None):
    """Delete a sample plus dependent records and clear staged import references."""
    documents = db.query(Document).filter(Document.sample_id == sample.id).all()
    deleted_files = 0
    for doc in documents:
        if _safe_delete_upload_file(doc.file_path):
            deleted_files += 1

    counts = {
        "documents": len(documents),
        "tastings": db.query(Tasting).filter(Tasting.sample_id == sample.id).count(),
        "shipments": db.query(Shipment).filter(Shipment.sample_id == sample.id).count(),
        "events": db.query(Event).filter(Event.sample_id == sample.id).count(),
        "files": deleted_files,
    }

    import_rows_query = db.query(ImportRow).filter(ImportRow.sample_id == sample.id)
    if import_batch_id is not None:
        import_rows_query = import_rows_query.filter(ImportRow.batch_id == import_batch_id)
    import_rows = import_rows_query.all()
    for row in import_rows:
        row.sample_id = None
        if import_final_action:
            row.final_action = import_final_action
            row.status = "deleted"

    db.query(Document).filter(Document.sample_id == sample.id).delete(synchronize_session=False)
    db.query(Event).filter(Event.sample_id == sample.id).delete(synchronize_session=False)
    db.query(Shipment).filter(Shipment.sample_id == sample.id).delete(synchronize_session=False)
    db.query(Tasting).filter(Tasting.sample_id == sample.id).delete(synchronize_session=False)
    db.delete(sample)
    return counts


@app.on_event("startup")
async def startup_event():
    """Startup event"""
    # Create necessary directories
    Path("./data").mkdir(exist_ok=True)
    uploads_path.mkdir(parents=True, exist_ok=True)


# ============================================================================
# DASHBOARD AND MAIN VIEWS
# ============================================================================

@app.get("/", response_class=HTMLResponse)
async def dashboard(request: Request, db: Session = Depends(get_db)):
    """Dashboard principal"""
    total_samples = db.query(Sample).count()
    received_samples = db.query(Sample).filter(
        Sample.status == SampleStatus.RECEIVED
    ).count()
    available_samples = db.query(Sample).filter(
        Sample.status == SampleStatus.AVAILABLE
    ).count()
    approved_samples = db.query(Sample).filter(
        Sample.status == SampleStatus.APPROVED
    ).count()
    rejected_samples = db.query(Sample).filter(
        Sample.status == SampleStatus.REJECTED
    ).count()
    shipped_samples = db.query(Sample).filter(
        Sample.status == SampleStatus.SHIPPED
    ).count()
    archived_samples = db.query(Sample).filter(
        Sample.status == SampleStatus.ARCHIVED
    ).count()
    out_of_stock = db.query(Sample).filter(
        Sample.available_quantity_g <= 0
    ).count()
    # total_quantity in grams using available_quantity_g
    total_quantity = db.query(Sample).filter(
        Sample.available_quantity_g > 0
    ).with_entities(func.sum(Sample.available_quantity_g)).scalar() or 0
    
    recent_samples = db.query(Sample).order_by(desc(Sample.created_at)).limit(6).all()
    high_score_samples = db.query(Sample).join(Tasting).order_by(
        desc(Tasting.indian_score)
    ).limit(5).all()
    latest_tastings = db.query(Tasting).order_by(desc(Tasting.tasting_date)).limit(5).all()
    pending_tasting_samples = db.query(Sample).filter(~Sample.tastings.any()).order_by(desc(Sample.created_at)).limit(6).all()
    out_of_stock_samples = db.query(Sample).filter(Sample.available_quantity_g <= 0).order_by(desc(Sample.updated_at)).limit(6).all()
    origin_counts = db.query(Sample.origin, func.count(Sample.id)).group_by(Sample.origin).order_by(desc(func.count(Sample.id))).limit(6).all()
    provider_counts = db.query(Sample.producer, func.count(Sample.id)).group_by(Sample.producer).order_by(desc(func.count(Sample.id))).limit(6).all()
    
    return templates.TemplateResponse("dashboard.html", {
        "request": request,
        "total_samples": total_samples,
        "received_samples": received_samples,
        "available_samples": available_samples,
        "approved_samples": approved_samples,
        "rejected_samples": rejected_samples,
        "shipped_samples": shipped_samples,
        "archived_samples": archived_samples,
        "out_of_stock": out_of_stock,
        "total_quantity": total_quantity,
        "recent_samples": recent_samples,
        "high_score_samples": high_score_samples,
        "latest_tastings": latest_tastings,
        "pending_tasting_samples": pending_tasting_samples,
        "out_of_stock_samples": out_of_stock_samples,
        "origin_counts": origin_counts,
        "provider_counts": provider_counts,
    })


# ============================================================================
# SAMPLE MANAGEMENT
# ============================================================================

@app.get("/samples", response_class=HTMLResponse)
async def list_samples(
    request: Request,
    code: str = None,
    status: str = None,
    country: str = None,
    producer: str = None,
    supplier_reference: str = None,
    origin: str = None,
    purchase_cvc: str = None,
    sales_cvv: str = None,
    quality: str = None,
    commercial_result: str = None,
    indian_min: float = None,
    indian_max: float = None,
    message: str = None,
    db: Session = Depends(get_db)
):
    """Lista de muestras"""
    query = db.query(Sample)
    
    if code:
        query = query.filter(Sample.code.ilike(f"%{code}%"))
    if status:
        normalized_status = normalize_status_value(status)
        if normalized_status in {s.value for s in SampleStatus}:
            query = query.filter(Sample.status == normalized_status)
    if country:
        query = query.filter(or_(
            Sample.country_code == country,
            Sample.country_name.ilike(f"%{country}%"),
            Sample.origin.ilike(f"%{country}%"),
        ))
    if producer:
        query = query.filter(Sample.producer.ilike(f"%{producer}%"))
    if origin:
        query = query.filter(Sample.origin.ilike(f"%{origin}%"))
    if supplier_reference:
        query = query.filter(Sample.supplier_reference.ilike(f"%{supplier_reference}%"))
    if purchase_cvc:
        query = query.filter(Sample.purchase_contract_cvc.ilike(f"%{purchase_cvc}%"))
    if sales_cvv:
        query = query.filter(Sample.sales_contract_cvv == sales_cvv)
    if quality:
        query = query.filter(Sample.quality.ilike(f"%{quality}%"))
    if commercial_result:
        query = query.filter(Sample.commercial_result == commercial_result)

    # Indian score range filter: join tastings
    if indian_min is not None or indian_max is not None:
        tquery = db.query(Sample).join(Tasting)
        if indian_min is not None:
            tquery = tquery.filter(Tasting.indian_score >= indian_min)
        if indian_max is not None:
            tquery = tquery.filter(Tasting.indian_score <= indian_max)
        ids = [s.id for s in tquery.distinct().all()]
        query = query.filter(Sample.id.in_(ids))
    
    samples = query.order_by(desc(Sample.created_at)).all()
    
    return templates.TemplateResponse("samples.html", {
        "request": request,
        "samples": samples,
        "country_options": _sample_country_options(db),
        "provider_options": _sample_field_options(db, "producer"),
        "quality_options": _sample_field_options(db, "quality"),
        "origin_options": _sample_field_options(db, "origin"),
        "statuses": [{"value": s.value, "label": status_label(s.value)} for s in SampleStatus],
        "message": message,
        "filters": {
            "code": code or "",
            "status": status or "",
            "country": country or "",
            "producer": producer or "",
            "supplier_reference": supplier_reference or "",
            "origin": origin or "",
            "purchase_cvc": purchase_cvc or "",
            "quality": quality or "",
        },
    })


@app.get("/samples/new", response_class=HTMLResponse)
async def new_sample_form(request: Request):
    """Formulario para nueva muestra"""
    return templates.TemplateResponse("sample_form.html", {
        "request": request,
        "countries": get_all_countries(),
        "processing_methods": [
            "Washed/Lavado", "Natural/Secado al sol", "Honey/Miel",
            "Anaerobic", "Fermented"
        ]
    })


@app.post("/samples")
async def create_sample(
    code: str = Form(None),
    producer: str = Form(...),
    supplier_reference: str = Form(...),
    quality: str = Form(...),
    received_date: str = Form(None),
    initial_quantity_grams: str = Form(None),
    available_quantity_grams: str = Form(None),
    country_code: str = Form(None),
    origin: str = Form(None),
    provider_sample_number: str = Form(None),
    container_number: str = Form(None),
    purchase_contract_cvc: str = Form(None),
    sales_contract_cvv: str = Form(None),
    warehouse: str = Form(None),
    sample_type: str = Form(None),
    category: str = Form(None),
    commercial_result: str = Form(None),
    harvest_date: str = Form(None),
    variety: str = Form(None),
    altitude: int = Form(None),
    processing: str = Form(None),
    physical_location: str = Form(None),
    notes: str = Form(None),
    db: Session = Depends(get_db)
):
    """Crear nueva muestra"""
    # Auto-generate code if not provided
    if not code or code.strip() == '':
        import random
        import string
        code = f"AUTO-{datetime.utcnow().strftime('%Y%m%d')}-{''.join(random.choices(string.ascii_uppercase + string.digits, k=4))}"
    
    existing = db.query(Sample).filter(Sample.code == code).first()
    if existing:
        raise HTTPException(status_code=400, detail="El código de muestra ya existe")

    try:
        initial_quantity = float(initial_quantity_grams) / 1000.0 if initial_quantity_grams not in (None, '', '0') else 0.0
    except ValueError:
        initial_quantity = 0.0
    try:
        available_quantity = float(available_quantity_grams) / 1000.0 if available_quantity_grams not in (None, '') else initial_quantity
    except ValueError:
        available_quantity = initial_quantity

    if initial_quantity < 0 or available_quantity < 0:
        raise HTTPException(status_code=400, detail="Las cantidades no pueden ser negativas")
    if initial_quantity != 0 and available_quantity > initial_quantity:
        raise HTTPException(status_code=400, detail="La cantidad disponible no puede ser mayor que la cantidad recibida")

    created_at = datetime.utcnow()
    if received_date:
        try:
            created_at = datetime.fromisoformat(received_date)
        except ValueError:
            created_at = datetime.utcnow()

    received_quantity_g = 0
    try:
        received_quantity_g = int(initial_quantity_grams) if initial_quantity_grams not in (None, '') else 0
    except (ValueError, TypeError):
        received_quantity_g = 0

    available_quantity_g = received_quantity_g
    try:
        if available_quantity_grams not in (None, ''):
            available_quantity_g = int(available_quantity_grams)
    except (ValueError, TypeError):
        available_quantity_g = received_quantity_g

    sample = Sample(
        code=code,
        country_code=country_code,
        country_name=get_country_name(country_code) if country_code else None,
        origin=origin,
        producer=producer,
        supplier_reference=supplier_reference,
        provider_sample_number=provider_sample_number,
        container_number=container_number,
        purchase_contract_cvc=purchase_contract_cvc,
        sales_contract_cvv=sales_contract_cvv,
        quality=quality,
        warehouse=warehouse,
        sample_type=sample_type,
        category=category,
        commercial_result=commercial_result,
        harvest_date=harvest_date,
        variety=variety,
        altitude=altitude,
        processing=processing,
        physical_location=physical_location,
        initial_quantity=initial_quantity,
        available_quantity=available_quantity,
        received_quantity_g=received_quantity_g,
        available_quantity_g=available_quantity_g,
        notes=notes,
        status=SampleStatus.RECEIVED,
        created_at=created_at
    )

    db.add(sample)
    db.commit()
    db.refresh(sample)

    event = Event(
        sample_id=sample.id,
        event_type="received",
        description=f"Muestra registrada - {received_quantity_g} g"
    )
    db.add(event)
    db.commit()

    return JSONResponse({"id": sample.id, "code": sample.code})


@app.post("/samples/delete-selected")
async def delete_selected_samples(request: Request, db: Session = Depends(get_db)):
    form = await request.form()
    sample_ids = []
    for value in form.getlist("sample_ids"):
        try:
            sample_ids.append(int(value))
        except (TypeError, ValueError):
            continue
    if not sample_ids:
        message = quote_plus("Selecciona al menos una muestra para eliminar")
        return RedirectResponse(f"/samples?message={message}", status_code=303)

    deleted = []
    failed = []
    for sample_id in sample_ids:
        sample = db.query(Sample).filter(Sample.id == sample_id).first()
        if not sample:
            failed.append(f"#{sample_id}: no encontrada")
            continue
        try:
            sample_label = sample.code or f"#{sample.id}"
            _delete_sample_records(db, sample, import_final_action="DELETED_MANUAL")
            deleted.append(sample_label)
        except Exception as exc:
            db.rollback()
            failed.append(f"#{sample_id}: {exc}")

    db.commit()
    summary = f"Eliminadas correctamente: {len(deleted)}"
    if failed:
        summary += f". No eliminadas: {len(failed)} ({'; '.join(failed[:3])})"
        if len(failed) > 3:
            summary += "..."
    message = quote_plus(summary)
    return RedirectResponse(f"/samples?message={message}", status_code=303)


@app.get("/samples/{sample_id}/edit", response_class=HTMLResponse)
async def edit_sample_form(sample_id: int, request: Request, db: Session = Depends(get_db)):
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")

    return templates.TemplateResponse("sample_form.html", {
        "request": request,
        "sample": sample,
        "action_url": f"/samples/{sample.id}",
        "submit_label": "Actualizar muestra",
        "countries": get_all_countries(),
        "processing_methods": [
            "Washed/Lavado", "Natural/Secado al sol", "Honey/Miel",
            "Anaerobic", "Fermented"
        ]
    })


@app.api_route("/samples/{sample_id}", methods=["POST", "PUT"])
async def update_sample(
    sample_id: int,
    code: str = Form(None),
    producer: str = Form(...),
    supplier_reference: str = Form(...),
    quality: str = Form(...),
    received_date: str = Form(None),
    initial_quantity_grams: int = Form(None),
    available_quantity_grams: int = Form(None),
    country_code: str = Form(None),
    origin: str = Form(None),
    provider_sample_number: str = Form(None),
    container_number: str = Form(None),
    purchase_contract_cvc: str = Form(None),
    sales_contract_cvv: str = Form(None),
    warehouse: str = Form(None),
    sample_type: str = Form(None),
    category: str = Form(None),
    commercial_result: str = Form(None),
    harvest_date: str = Form(None),
    variety: str = Form(None),
    altitude: int = Form(None),
    processing: str = Form(None),
    physical_location: str = Form(None),
    notes: str = Form(None),
    db: Session = Depends(get_db)
):
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")
    
    # If code not provided, keep existing
    if not code or code.strip() == '':
        code = sample.code
    
    existing = db.query(Sample).filter(Sample.code == code, Sample.id != sample_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="El código de muestra ya existe")

    try:
        initial_quantity = float(initial_quantity_grams) / 1000.0 if initial_quantity_grams not in (None, '', '0') else 0.0
    except ValueError:
        initial_quantity = 0.0
    try:
        available_quantity = float(available_quantity_grams) / 1000.0 if available_quantity_grams not in (None, '') else initial_quantity
    except ValueError:
        available_quantity = initial_quantity

    if initial_quantity < 0 or available_quantity < 0:
        raise HTTPException(status_code=400, detail="Las cantidades no pueden ser negativas")
    if initial_quantity != 0 and available_quantity > initial_quantity:
        raise HTTPException(status_code=400, detail="La cantidad disponible no puede ser mayor que la cantidad recibida")

    sample.code = code
    sample.country_code = country_code
    sample.country_name = get_country_name(country_code) if country_code else None
    sample.origin = origin
    sample.producer = producer
    sample.supplier_reference = supplier_reference
    sample.provider_sample_number = provider_sample_number
    sample.container_number = container_number
    sample.purchase_contract_cvc = purchase_contract_cvc
    sample.sales_contract_cvv = sales_contract_cvv
    sample.quality = quality
    sample.warehouse = warehouse
    sample.sample_type = sample_type
    sample.category = category
    sample.commercial_result = commercial_result
    sample.harvest_date = harvest_date
    sample.variety = variety
    sample.altitude = altitude
    sample.processing = processing
    sample.physical_location = physical_location
    sample.initial_quantity = initial_quantity
    sample.available_quantity = available_quantity
    try:
        sample.received_quantity_g = int(initial_quantity_grams) if initial_quantity_grams not in (None, '') else sample.received_quantity_g
    except (ValueError, TypeError):
        pass
    try:
        sample.available_quantity_g = int(available_quantity_grams) if available_quantity_grams not in (None, '') else sample.received_quantity_g
    except (ValueError, TypeError):
        pass
    sample.notes = notes

    if received_date:
        try:
            sample.created_at = datetime.fromisoformat(received_date)
        except ValueError:
            pass

    db.commit()
    db.refresh(sample)
    return JSONResponse({"id": sample.id, "code": sample.code})


@app.get("/samples/{sample_id}", response_class=HTMLResponse)
async def sample_detail(
    sample_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    """Detalle de muestra"""
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")
    
    tastings = db.query(Tasting).filter(Tasting.sample_id == sample_id).all()
    shipments = db.query(Shipment).filter(Shipment.sample_id == sample_id).all()
    events = db.query(Event).filter(Event.sample_id == sample_id).order_by(
        desc(Event.event_date)
    ).all()
    documents = db.query(Document).filter(Document.sample_id == sample_id).all()
    
    best_tasting = None
    if tastings:
        best_tasting = max(tastings, key=lambda x: x.indian_score or 0)
    
    return templates.TemplateResponse("sample_detail.html", {
        "request": request,
        "sample": sample,
        "tastings": tastings,
        "shipments": shipments,
        "events": events,
        "documents": documents,
        "best_tasting": best_tasting,
        "flag": get_country_flag(sample.country_code),
    })


@app.post("/samples/{sample_id}/delete")
async def delete_sample(sample_id: int, db: Session = Depends(get_db)):
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")

    sample_code = sample.code or f"#{sample.id}"
    _delete_sample_records(db, sample, import_final_action="DELETED_MANUAL")
    db.commit()
    message = quote_plus(f"Muestra {sample_code} eliminada correctamente")
    return RedirectResponse(f"/samples?message={message}", status_code=303)


@app.post("/samples/{sample_id}/tastings/{tasting_id}/documents")
async def upload_tasting_document(sample_id: int, tasting_id: int, file: UploadFile = File(...), document_type: str = Form(None), db: Session = Depends(get_db)):
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    tasting = db.query(Tasting).filter(Tasting.id == tasting_id, Tasting.sample_id == sample_id).first()
    if not sample or not tasting:
        raise HTTPException(status_code=404, detail="Sample or tasting not found")

    upload_dir = os.path.join(UPLOADS_DIR, "samples", str(sample_id), "tastings", str(tasting_id))
    os.makedirs(upload_dir, exist_ok=True)
    filename = os.path.basename(file.filename)
    safe_name = f"{int(datetime.utcnow().timestamp())}_{filename}"
    save_path = os.path.join(upload_dir, safe_name)

    with open(save_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    doc = Document(sample_id=sample.id, tasting_id=tasting.id, file_name=filename, file_path=save_path, file_type=document_type or file.content_type)
    db.add(doc)
    db.commit()

    return RedirectResponse(f"/samples/{sample_id}", status_code=303)


def _pdf_value(value, default="-"):
    return default if value in (None, "") else str(value)


def _wrap_text(text, max_chars=78):
    import textwrap
    if not text:
        return []
    lines = []
    for paragraph in str(text).splitlines():
        lines.extend(textwrap.wrap(paragraph, width=max_chars) or [""])
    return lines


def _draw_section_title(c, x, y, title, accent="#B07A42"):
    c.setFillColor(HexColor(accent))
    c.roundRect(x, y - 5 * mm, 3 * mm, 5 * mm, 1.2 * mm, stroke=0, fill=1)
    c.setFillColor(HexColor("#153F2B"))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(x + 5 * mm, y - 3.7 * mm, title)


def _draw_field_grid(c, x, y, fields, col_width, row_height=10 * mm, cols=2):
    c.setStrokeColor(HexColor("#D7DEE3"))
    c.setLineWidth(0.35)
    for index, (label, value) in enumerate(fields):
        col = index % cols
        row = index // cols
        bx = x + col * col_width
        by = y - row * row_height
        c.setFillColor(HexColor("#F7FAF8"))
        c.roundRect(bx, by - row_height + 1 * mm, col_width - 3 * mm, row_height - 1.5 * mm, 2 * mm, stroke=1, fill=1)
        c.setFillColor(HexColor("#65756B"))
        c.setFont("Helvetica-Bold", 6.6)
        c.drawString(bx + 3 * mm, by - 3.3 * mm, label.upper())
        c.setFillColor(HexColor("#111827"))
        c.setFont("Helvetica", 8.2)
        display = _pdf_value(value)
        if len(display) > 34:
            display = display[:31] + "..."
        c.drawString(bx + 3 * mm, by - 7.2 * mm, display)
    return y - (((len(fields) + cols - 1) // cols) * row_height)


def _draw_score_tile(c, x, y, label, value, width=28 * mm, height=18 * mm):
    c.setFillColor(HexColor("#153F2B"))
    c.roundRect(x, y - height, width, height, 3 * mm, stroke=0, fill=1)
    c.setFillColor(HexColor("#D9B45E"))
    c.setFont("Helvetica-Bold", 6.5)
    c.drawCentredString(x + width / 2, y - 5 * mm, label.upper())
    c.setFillColor(HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 16)
    c.drawCentredString(x + width / 2, y - 13 * mm, _pdf_value(value))


def _draw_notes_box(c, x, y, width, title, text, max_lines=8):
    c.setStrokeColor(HexColor("#D7DEE3"))
    c.setFillColor(HexColor("#FFFFFF"))
    box_height = 10 * mm + max_lines * 4.2 * mm
    c.roundRect(x, y - box_height, width, box_height, 2.5 * mm, stroke=1, fill=1)
    c.setFillColor(HexColor("#153F2B"))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x + 4 * mm, y - 5 * mm, title)
    c.setFillColor(HexColor("#333333"))
    c.setFont("Helvetica", 8)
    text_y = y - 10 * mm
    for line in _wrap_text(text, 92)[:max_lines]:
        c.drawString(x + 4 * mm, text_y, line)
        text_y -= 4.2 * mm
    return y - box_height


def _draw_photo_slot(c, doc, x, y, width, height, title):
    c.setStrokeColor(HexColor("#D7DEE3"))
    c.setFillColor(HexColor("#F7FAF8"))
    c.roundRect(x, y, width, height, 2.5 * mm, stroke=1, fill=1)
    c.setFillColor(HexColor("#153F2B"))
    c.setFont("Helvetica-Bold", 8)
    c.drawString(x + 3 * mm, y + height - 5 * mm, title)
    if doc:
        try:
            c.drawImage(ImageReader(doc.file_path), x + 3 * mm, y + 4 * mm, width=width - 6 * mm, height=height - 12 * mm, preserveAspectRatio=True, anchor='c')
            return
        except Exception:
            pass
    c.setFillColor(HexColor("#9AA5AD"))
    c.setFont("Helvetica", 8)
    c.drawCentredString(x + width / 2, y + height / 2, "Sin fotografia")


def _render_tasting_pdf_premium(sample: Sample, tasting: Tasting, docs, sample_docs) -> io.BytesIO:
    photos = [doc for doc in docs if doc.file_type and ('verde' in doc.file_type.lower() or 'green' in doc.file_type.lower())]
    roast_photos = [doc for doc in docs if doc.file_type and ('tostado' in doc.file_type.lower() or 'roasted' in doc.file_type.lower())]
    if not photos or not roast_photos:
        fallback = [doc for doc in sample_docs if doc.file_type and doc.file_type.startswith('image')]
        if not photos and fallback:
            photos = [fallback[0]]
        if not roast_photos and len(fallback) > 1:
            roast_photos = [fallback[1]]

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4
    margin = 13 * mm
    primary = "#153F2B"
    gold = "#D9B45E"

    c.setFillColor(HexColor(primary))
    c.rect(0, height - 34 * mm, width, 34 * mm, stroke=0, fill=1)
    c.setFillColor(HexColor(gold))
    c.rect(0, height - 36 * mm, width, 2 * mm, stroke=0, fill=1)
    try:
        logo_path = Path(__file__).parent / "static" / "logo.png"
        if logo_path.exists():
            c.drawImage(ImageReader(str(logo_path)), margin, height - 27 * mm, width=44 * mm, height=16 * mm, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass
    c.setFillColor(HexColor("#FFFFFF"))
    c.setFont("Helvetica-Bold", 17)
    c.drawRightString(width - margin, height - 17 * mm, "Ficha tecnica de cata")
    c.setFont("Helvetica", 9)
    c.drawRightString(width - margin, height - 24 * mm, f"{sample.code} | {sample.country_name or '-'} | {sample.origin or '-'}")

    y = height - 45 * mm
    _draw_score_tile(c, margin, y, "Indian", f"{tasting.indian_score or 0:.1f}")
    _draw_score_tile(c, margin + 32 * mm, y, "Taza", f"{tasting.cup_score or 0:.1f}")
    _draw_score_tile(c, margin + 64 * mm, y, "Comercial", f"{tasting.commercial_score or 0:.1f}")
    _draw_score_tile(c, margin + 96 * mm, y, "Resultado", tasting_result_label(tasting.result), width=36 * mm)
    c.setFillColor(HexColor("#65756B"))
    c.setFont("Helvetica", 8)
    c.drawRightString(width - margin, y - 6 * mm, f"Fecha cata: {tasting.tasting_date.strftime('%d/%m/%Y') if tasting.tasting_date else '-'}")
    if tasting.roast_date:
        c.drawRightString(width - margin, y - 11 * mm, f"Fecha tueste: {tasting.roast_date.strftime('%d/%m/%Y')}")

    y -= 27 * mm
    _draw_section_title(c, margin, y, "Datos de muestra")
    y -= 9 * mm
    sample_fields = [
        ("Codigo", sample.code),
        ("Pais / origen", f"{sample.country_name or '-'} / {sample.origin or '-'}"),
        ("Proveedor", sample.producer),
        ("Ref. proveedor", sample.supplier_reference),
        ("Calidad", sample.quality),
        ("Variedad", sample.variety),
        ("Proceso", sample.processing),
        ("Cosecha", sample.harvest_date),
        ("Contenedor", sample.container_number),
        ("Contrato CVC", sample.purchase_contract_cvc),
    ]
    y = _draw_field_grid(c, margin, y, sample_fields, (width - margin * 2) / 2)

    y -= 6 * mm
    _draw_section_title(c, margin, y, "Datos fisicos")
    y -= 9 * mm
    physical_fields = [
        ("Humedad", f"{tasting.humidity or 0:.1f}%"),
        ("Defectos prim./sec.", f"{tasting.defects_primary or 0}/{tasting.defects_secondary or 0}"),
        ("Criba 18+", f"{tasting.sieve_18 or 0:.1f}%"),
        ("Criba 17", f"{tasting.sieve_17 or 0:.1f}%"),
        ("Criba 16+", f"{tasting.sieve_16 or 0:.1f}%"),
        ("Criba 15", f"{tasting.sieve_15 or 0:.1f}%"),
        ("Criba 14+", f"{tasting.sieve_14 or 0:.1f}%"),
        ("Criba plato", f"{tasting.sieve_plato or 0:.1f}%"),
    ]
    y = _draw_field_grid(c, margin, y, physical_fields, (width - margin * 2) / 4, cols=4)

    y -= 6 * mm
    _draw_section_title(c, margin, y, "Datos sensoriales")
    y -= 9 * mm
    sensory_fields = [
        ("Aroma", tasting.aroma),
        ("Acidez", tasting.acidity),
        ("Cuerpo", tasting.body),
        ("Sabor", tasting.flavor),
        ("Postgusto", tasting.aftertaste),
        ("Limpieza", tasting.cleanliness),
        ("Balance", tasting.balance),
        ("Valoracion", tasting.valuation),
    ]
    y = _draw_field_grid(c, margin, y, sensory_fields, (width - margin * 2) / 4, cols=4)

    y -= 6 * mm
    y = _draw_notes_box(c, margin, y, width - margin * 2, "Notas de cata", tasting.tasting_notes or "Sin notas de cata registradas.", max_lines=7)
    y -= 6 * mm
    y = _draw_notes_box(c, margin, y, width - margin * 2, "Recomendaciones", tasting.recommendations or "Sin recomendaciones registradas.", max_lines=4)

    photo_y = margin + 8 * mm
    photo_width = (width - margin * 2 - 8 * mm) / 2
    photo_height = max(40 * mm, y - photo_y - 5 * mm)
    _draw_photo_slot(c, photos[0] if photos else None, margin, photo_y, photo_width, photo_height, "Cafe verde")
    _draw_photo_slot(c, roast_photos[0] if roast_photos else None, margin + photo_width + 8 * mm, photo_y, photo_width, photo_height, "Cafe tostado")

    c.setFillColor(HexColor("#65756B"))
    c.setFont("Helvetica", 7)
    c.drawCentredString(width / 2, 7 * mm, "Indian Ecotrade | Control de muestras de cafe verde")
    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer


@app.get("/samples/{sample_id}/tastings/{tasting_id}/pdf")
async def tasting_pdf(sample_id: int, tasting_id: int, request: Request, db: Session = Depends(get_db)):
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    tasting = db.query(Tasting).filter(Tasting.id == tasting_id, Tasting.sample_id == sample_id).first()
    if not sample or not tasting:
        raise HTTPException(status_code=404, detail="Sample or tasting not found")

    docs = db.query(Document).filter(Document.tasting_id == tasting_id).all()
    sample_docs = db.query(Document).filter(Document.sample_id == sample_id, Document.tasting_id == None).all()

    pdf_buffer = _render_tasting_pdf_premium(sample, tasting, docs, sample_docs)
    headers = {"Content-Disposition": f"attachment; filename=\"ficha_cata_{sample.code}_{tasting_id}.pdf\""}
    return StreamingResponse(pdf_buffer, media_type="application/pdf", headers=headers)

    pdf_buffer = io.BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=A4)
    width, height = A4
    margin = 15 * mm

    # Header with logo
    try:
        logo_path = Path(__file__).parent / "static" / "logo.png"
        if logo_path.exists():
            c.drawImage(ImageReader(str(logo_path)), margin, height - 35*mm, width=45*mm, height=20*mm, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    c.setFont("Helvetica-Bold", 16)
    c.drawString(margin + 50*mm, height - 28*mm, "Ficha de cata")
    c.setFont("Helvetica", 10)
    c.drawString(margin + 50*mm, height - 35*mm, f"Muestra: {sample.code}")

    # Sample information
    y = height - 45*mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(margin, y, "Datos de la muestra")
    y -= 6*mm
    c.setFont("Helvetica", 9)
    c.drawString(margin, y, f"Origen: {sample.origin or '-'}")
    y -= 5*mm
    c.drawString(margin, y, f"País: {sample.country_name or '-'}")
    y -= 5*mm
    c.drawString(margin, y, f"Calidad: {sample.quality or '-'}")
    y -= 5*mm
    c.drawString(margin, y, f"Proveedor: {sample.producer or '-'}")
    y -= 5*mm
    c.drawString(margin, y, f"Ref. proveedor: {sample.supplier_reference or '-'}")
    y -= 5*mm
    if sample.container_number:
        c.drawString(margin, y, f"Contenedor: {sample.container_number}")
        y -= 5*mm
    if sample.purchase_contract_cvc:
        c.drawString(margin, y, f"CVC: {sample.purchase_contract_cvc}")
        y -= 5*mm

    # Tasting physical data
    y -= 4*mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(margin, y, "Datos físicos de cata")
    y -= 6*mm
    c.setFont("Helvetica", 9)
    c.drawString(margin, y, f"Fecha: {tasting.tasting_date.strftime('%Y-%m-%d')}")
    y -= 5*mm
    if tasting.roast_date:
        c.drawString(margin, y, f"Fecha tostado: {tasting.roast_date.strftime('%Y-%m-%d')}")
        y -= 5*mm
    c.drawString(margin, y, f"Humedad: {tasting.humidity or 0:.1f}%")
    y -= 5*mm
    c.drawString(margin, y, f"Criba 18+: {tasting.sieve_18 or 0:.1f}%  16+: {tasting.sieve_16 or 0:.1f}%")
    y -= 5*mm
    c.drawString(margin, y, f"14+: {tasting.sieve_14 or 0:.1f}%")
    y -= 5*mm
    c.drawString(margin, y, f"Defectos prim./sec.: {tasting.defects_primary or 0}/{tasting.defects_secondary or 0}")
    y -= 5*mm
    c.drawString(margin, y, f"Resultado de cata: {tasting_result_label(tasting.result)}")

    # Sensory data
    y -= 8*mm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(margin, y, "Datos sensoriales")
    y -= 6*mm
    c.setFont("Helvetica", 9)
    c.drawString(margin, y, f"Aroma: {tasting.aroma or 0}  Acidez: {tasting.acidity or 0}  Cuerpo: {tasting.body or 0}")
    y -= 5*mm
    c.drawString(margin, y, f"Sabor: {tasting.flavor or 0}  Postgusto: {tasting.aftertaste or 0}  Limpieza: {tasting.cleanliness or 0}")
    y -= 5*mm
    c.drawString(margin, y, f"Balance: {tasting.balance or 0}  Cup Score: {tasting.cup_score or 0:.1f}  Indian: {tasting.indian_score or 0:.1f}")
    y -= 8*mm

    # Notes
    if tasting.tasting_notes:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(margin, y, "Notas")
        y -= 6*mm
        c.setFont("Helvetica", 9)
        text = c.beginText(margin, y)
        text.setLeading(10)
        for line in tasting.tasting_notes.splitlines():
            text.textLine(line)
        c.drawText(text)
        y = text.getY() - 8*mm
    else:
        y -= 4*mm

    # Photos
    photos = [doc for doc in docs if doc.file_type and ('verde' in doc.file_type.lower() or 'green' in doc.file_type.lower())]
    roast_photos = [doc for doc in docs if doc.file_type and ('tostado' in doc.file_type.lower() or 'roasted' in doc.file_type.lower())]
    if not photos or not roast_photos:
        fallback = [doc for doc in sample_docs if doc.file_type and doc.file_type.startswith('image')]
        if photos == [] and fallback:
            photos = [fallback[0]]
        if roast_photos == [] and len(fallback) > 1:
            roast_photos = [fallback[1]]

    img_width = (width - margin * 2 - 10*mm) / 2
    img_height = 70*mm
    row_y = margin + img_height
    x_left = margin
    x_right = margin + img_width + 10*mm

    if photos:
        try:
            c.drawImage(ImageReader(photos[0].file_path), x_left, margin, width=img_width, height=img_height, preserveAspectRatio=True, anchor='sw')
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x_left, margin + img_height + 2*mm, "Café verde")
        except Exception:
            pass
    if roast_photos:
        try:
            c.drawImage(ImageReader(roast_photos[0].file_path), x_right, margin, width=img_width, height=img_height, preserveAspectRatio=True, anchor='sw')
            c.setFont("Helvetica-Bold", 9)
            c.drawString(x_right, margin + img_height + 2*mm, "Café tostado")
        except Exception:
            pass

    c.save()
    pdf_buffer.seek(0)
    headers = {"Content-Disposition": f"attachment; filename=\"ficha_cata_{sample.code}_{tasting_id}.pdf\""}
    return StreamingResponse(pdf_buffer, media_type="application/pdf", headers=headers)


@app.post("/samples/{sample_id}/documents")
async def upload_document(sample_id: int, file: UploadFile = File(...), document_type: str = Form(None), db: Session = Depends(get_db)):
    """Upload a document or photo for a sample"""
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Sample not found")

    upload_dir = os.path.join(UPLOADS_DIR, str(sample_id))
    os.makedirs(upload_dir, exist_ok=True)
    filename = os.path.basename(file.filename)
    safe_name = f"{int(datetime.utcnow().timestamp())}_{filename}"
    save_path = os.path.join(upload_dir, safe_name)

    with open(save_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    doc = Document(sample_id=sample.id, file_name=filename, file_path=save_path, file_type=document_type or file.content_type)
    db.add(doc)
    db.commit()

    return RedirectResponse(f"/samples/{sample_id}", status_code=303)


# ============================================================================
# TASTING/CUPPING
# ============================================================================

def _optional_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _optional_int(value, default=0):
    if value in (None, ""):
        return default
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _average_present(values, default=0.0):
    present = [value for value in values if value is not None]
    if not present:
        return default
    return sum(present) / len(present)


@app.post("/samples/{sample_id}/tastings")
async def create_tasting(
    sample_id: int,
    evaluator: str = Form(...),
    tasting_date: str = Form(None),
    roast_date: str = Form(None),
    sieve_18: str = Form(None),
    sieve_17: str = Form(None),
    sieve_16: str = Form(None),
    sieve_15: str = Form(None),
    sieve_14: str = Form(None),
    sieve_13: str = Form(None),
    sieve_12: str = Form(None),
    sieve_plato: str = Form(None),
    humidity: str = Form(None),
    defects_primary: str = Form(None),
    defects_secondary: str = Form(None),
    aroma: str = Form(None),
    acidity: str = Form(None),
    body: str = Form(None),
    flavor: str = Form(None),
    aftertaste: str = Form(None),
    cleanliness: str = Form(None),
    balance: str = Form(None),
    tasting_notes: str = Form(None),
    recommendations: str = Form(None),
    valuation: str = Form(None),
    result: str = Form(None),
    redirect: str = Form(None),
    db: Session = Depends(get_db)
):
    """Crear cata/evaluación"""
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")

    evaluator = (evaluator or "").strip()
    if not evaluator:
        raise HTTPException(status_code=400, detail="El evaluador es obligatorio")
    
    # Parse optional dates
    td = datetime.utcnow()
    if tasting_date:
        try:
            td = datetime.fromisoformat(tasting_date)
        except Exception:
            td = datetime.utcnow()

    rd = None
    if roast_date:
        try:
            rd = datetime.fromisoformat(roast_date)
        except Exception:
            rd = None

    sieve_18_value = _optional_float(sieve_18)
    sieve_17_value = _optional_float(sieve_17)
    sieve_16_value = _optional_float(sieve_16)
    sieve_15_value = _optional_float(sieve_15)
    sieve_14_value = _optional_float(sieve_14)
    sieve_13_value = _optional_float(sieve_13)
    sieve_12_value = _optional_float(sieve_12)
    sieve_plato_value = _optional_float(sieve_plato)
    humidity_value = _optional_float(humidity)
    defects_primary_value = _optional_int(defects_primary, 0)
    defects_secondary_value = _optional_int(defects_secondary, 0)
    aroma_value = _optional_float(aroma)
    acidity_value = _optional_float(acidity)
    body_value = _optional_float(body)
    flavor_value = _optional_float(flavor)
    aftertaste_value = _optional_float(aftertaste)
    cleanliness_value = _optional_float(cleanliness)
    balance_value = _optional_float(balance)
    valuation_value = _optional_float(valuation)

    # Calculate scores
    cup_score = _average_present([
        aroma_value,
        acidity_value,
        body_value,
        flavor_value,
        aftertaste_value,
        cleanliness_value,
        balance_value,
    ])
    
    # Indian Score: proprietary calculation
    # SCA-based with custom adjustments
    base_score = cup_score
    sieve_bonus = (sieve_18_value * 0.5) if sieve_18_value is not None and sieve_18_value >= 50 else 0
    humidity_penalty = 5 if humidity_value is not None and humidity_value > 12 else 0
    defect_penalty = (defects_primary_value * 2) + defects_secondary_value
    
    indian_score = min(100, max(0, base_score + (sieve_bonus - humidity_penalty - defect_penalty)))
    
    # Commercial score based on commercial viability
    commercial_score = indian_score
    if humidity_value is not None and humidity_value > 12:
        commercial_score -= 10
    if defects_primary_value > 5:
        commercial_score -= 20
    
    commercial_score = max(0, min(100, commercial_score))
    
    tasting = Tasting(
        sample_id=sample_id,
        evaluator=evaluator,
        tasting_date=td,
        roast_date=rd,
        sieve_18=sieve_18_value,
        sieve_17=sieve_17_value,
        sieve_16=sieve_16_value,
        sieve_15=sieve_15_value,
        sieve_14=sieve_14_value,
        sieve_13=sieve_13_value,
        sieve_12=sieve_12_value,
        sieve_plato=sieve_plato_value,
        humidity=humidity_value,
        defects_primary=defects_primary_value,
        defects_secondary=defects_secondary_value,
        aroma=aroma_value,
        acidity=acidity_value,
        body=body_value,
        flavor=flavor_value,
        aftertaste=aftertaste_value,
        cleanliness=cleanliness_value,
        balance=balance_value,
        cup_score=cup_score,
        indian_score=indian_score,
        commercial_score=commercial_score,
        tasting_notes=tasting_notes,
        recommendations=recommendations,
        valuation=valuation_value,
        result=(result if result in ["pending","approved","rejected"] else "pending")
    )
    
    db.add(tasting)
    db.commit()
    db.refresh(tasting)

    # Update sample status using only the allowed sample states.
    if result == "approved":
        sample.status = SampleStatus.APPROVED
    elif result == "rejected":
        sample.status = SampleStatus.REJECTED
    elif (sample.available_quantity_g or 0) > 0 and sample.status == SampleStatus.RECEIVED:
        sample.status = SampleStatus.AVAILABLE

    # Create event
    event = Event(
        sample_id=sample_id,
        tasting_id=tasting.id,
        event_type="tasted",
        description=f"Evaluada con puntuación Indian Score: {indian_score:.1f}"
    )
    db.add(event)
    db.commit()

    if redirect:
        return JSONResponse({"id": sample.id, "tasting_id": tasting.id, "indian_score": indian_score})

    return JSONResponse({"id": tasting.id, "sample_id": sample.id, "indian_score": indian_score})


# ============================================================================
# SHIPMENTS
# ============================================================================

@app.post("/samples/{sample_id}/shipments")
async def create_shipment(
    sample_id: int,
    quantity_g: int = Form(...),
    destination: str = Form(...),
    reference: str = Form(...),
    notes: str = Form(None),
    db: Session = Depends(get_db)
):
    """Registrar envío"""
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")
    
    if quantity_g > sample.available_quantity_g:
        raise HTTPException(
            status_code=400,
            detail=f"Cantidad solicitada ({quantity_g} g) mayor que disponible ({sample.available_quantity_g} g)"
        )
    
    # Create shipment
    shipment = Shipment(
        sample_id=sample_id,
        quantity=quantity_g / 1000.0,
        quantity_g=quantity_g,
        destination=destination,
        reference=reference,
        notes=notes,
        status="pending"
    )
    
    # Reduce available quantity (grams and kg)
    sample.available_quantity_g -= quantity_g
    sample.available_quantity -= (quantity_g / 1000.0)
    
    # Update status
    if sample.available_quantity <= 0:
        sample.status = SampleStatus.SHIPPED
    elif sample.available_quantity < sample.initial_quantity:
        sample.status = SampleStatus.AVAILABLE
    
    # Create event
    event = Event(
        sample_id=sample_id,
        event_type="shipped",
        description=f"Envío de {quantity_g} g a {destination}"
    )
    
    db.add(shipment)
    db.add(event)
    db.commit()
    db.refresh(shipment)
    
    return JSONResponse({"id": shipment.id, "quantity_g": quantity_g})


# ============================================================================
# PDF LABELS WITH QR
# ============================================================================

AVERY_L7108REV = {
    "page_width_mm": 210,
    "page_height_mm": 297,
    "cell_width": 62 * mm,
    "cell_height": 89 * mm,
    "label_width": 89 * mm,
    "label_height": 62 * mm,
    "columns": 3,
    "rows": 3,
    "margin_x": 12 * mm,
    "margin_y": 15 * mm,
    "gap_x": 0,
    "gap_y": 0,
}


def _country_header(sample: Sample) -> str:
    flag = get_country_flag(sample.country_code) if sample.country_code else ""
    country = sample.country_name or sample.origin or "-"
    return f"{flag} {country}".strip()


def _normalize_origin_key(origin) -> str:
    text = unicodedata.normalize("NFKD", str(origin or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    return re.sub(r"\s+", " ", text)


def _quality_style_flags(quality) -> tuple[bool, bool]:
    key = _normalize_origin_key(quality)
    tokens = set(key.split())
    eco = bool(tokens.intersection({"eco", "ecologico", "organico", "organic", "bio"}))
    decaf = any(marker in key for marker in ["descafeinado", "decaf", "decaff", "decaffeinated", "desca"])
    return eco, decaf


def get_origin_flag_colors(origin) -> list[str]:
    if not origin:
        return []
    key = _normalize_origin_key(origin)
    aliases = {
        "peru": ["#D91023", "#FFFFFF", "#D91023"],
        "pe": ["#D91023", "#FFFFFF", "#D91023"],
        "colombia": ["#FCD116", "#003893", "#CE1126"],
        "co": ["#FCD116", "#003893", "#CE1126"],
        "brasil": ["#009B3A", "#FFDF00", "#002776"],
        "brazil": ["#009B3A", "#FFDF00", "#002776"],
        "br": ["#009B3A", "#FFDF00", "#002776"],
        "vietnam": ["#DA251D", "#FFFF00"],
        "viet nam": ["#DA251D", "#FFFF00"],
        "vn": ["#DA251D", "#FFFF00"],
        "uganda": ["#000000", "#FCDC04", "#D90000"],
        "ug": ["#000000", "#FCDC04", "#D90000"],
        "etiopia": ["#078930", "#FCDD09", "#DA121A"],
        "ethiopia": ["#078930", "#FCDD09", "#DA121A"],
        "et": ["#078930", "#FCDD09", "#DA121A"],
        "nicaragua": ["#0067C6", "#FFFFFF", "#0067C6"],
        "ni": ["#0067C6", "#FFFFFF", "#0067C6"],
        "honduras": ["#0073CF", "#FFFFFF", "#0073CF"],
        "hn": ["#0073CF", "#FFFFFF", "#0073CF"],
        "india": ["#FF9933", "#FFFFFF", "#138808"],
        "in": ["#FF9933", "#FFFFFF", "#138808"],
        "mexico": ["#006847", "#FFFFFF", "#CE1126"],
        "mx": ["#006847", "#FFFFFF", "#CE1126"],
        "guatemala": ["#4997D0", "#FFFFFF", "#4997D0"],
        "gt": ["#4997D0", "#FFFFFF", "#4997D0"],
        "costa rica": ["#002B7F", "#FFFFFF", "#CE1126", "#FFFFFF", "#002B7F"],
        "cr": ["#002B7F", "#FFFFFF", "#CE1126", "#FFFFFF", "#002B7F"],
        "el salvador": ["#0047AB", "#FFFFFF", "#0047AB"],
        "salvador": ["#0047AB", "#FFFFFF", "#0047AB"],
        "sv": ["#0047AB", "#FFFFFF", "#0047AB"],
        "kenia": ["#000000", "#BB0000", "#006600"],
        "kenya": ["#000000", "#BB0000", "#006600"],
        "ke": ["#000000", "#BB0000", "#006600"],
        "ruanda": ["#00A1DE", "#FAD201", "#20603D"],
        "rwanda": ["#00A1DE", "#FAD201", "#20603D"],
        "rw": ["#00A1DE", "#FAD201", "#20603D"],
        "burundi": ["#1EB53A", "#FFFFFF", "#CE1126"],
        "bi": ["#1EB53A", "#FFFFFF", "#CE1126"],
        "tanzania": ["#1EB53A", "#FCD116", "#000000", "#00A3DD"],
        "tz": ["#1EB53A", "#FCD116", "#000000", "#00A3DD"],
        "indonesia": ["#FF0000", "#FFFFFF"],
        "id": ["#FF0000", "#FFFFFF"],
        "costa de marfil": ["#F77F00", "#FFFFFF", "#009E60"],
        "cote d ivoire": ["#F77F00", "#FFFFFF", "#009E60"],
        "cote divoire": ["#F77F00", "#FFFFFF", "#009E60"],
        "ivory coast": ["#F77F00", "#FFFFFF", "#009E60"],
        "ci": ["#F77F00", "#FFFFFF", "#009E60"],
        "sierra leona": ["#1EB53A", "#FFFFFF", "#0072C6"],
        "sierra leone": ["#1EB53A", "#FFFFFF", "#0072C6"],
        "sl": ["#1EB53A", "#FFFFFF", "#0072C6"],
    }
    return aliases.get(key, [])


def _sample_flag_colors(sample: Sample) -> list[str]:
    return (
        get_origin_flag_colors(sample.country_code)
        or get_origin_flag_colors(sample.country_name)
        or get_origin_flag_colors(sample.origin)
    )


def _text_color_for_hex(hex_color: str) -> HexColor:
    color = hex_color.lstrip("#")
    try:
        r = int(color[0:2], 16)
        g = int(color[2:4], 16)
        b = int(color[4:6], 16)
    except Exception:
        return HexColor("#FFFFFF")
    luminance = (0.299 * r + 0.587 * g + 0.114 * b)
    return HexColor("#111827" if luminance > 170 else "#FFFFFF")


def _build_qr_image(data: str):
    qr = qrcode.QRCode(version=1, box_size=6, border=1)
    qr.add_data(data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_path = io.BytesIO()
    qr_img.save(qr_path, format="PNG")
    qr_path.seek(0)
    return ImageReader(qr_path)


def _draw_avery_l7108rev_label(c, sample: Sample, x: float, y: float, qr_data: str):
    width = AVERY_L7108REV["label_width"]
    height = AVERY_L7108REV["label_height"]
    padding = 4 * mm
    band_height = 10 * mm
    qr_size = 36 * mm

    c.saveState()
    c.setStrokeColor(HexColor("#D7DEE3"))
    c.setLineWidth(0.25)
    c.rect(x, y, width, height, stroke=1, fill=0)

    colors = _sample_flag_colors(sample)
    if colors:
        stripe_width = width / len(colors)
        for index, color in enumerate(colors):
            c.setFillColor(HexColor(color))
            c.rect(x + index * stripe_width, y + height - band_height, stripe_width, band_height, stroke=0, fill=1)
        title_bg = colors[0]
    else:
        title_bg = "#153F2B"
        c.setFillColor(HexColor(title_bg))
        c.rect(x, y + height - band_height, width, band_height, stroke=0, fill=1)

    c.setFillColor(_text_color_for_hex(title_bg))
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x + padding, y + height - 6.5 * mm, _country_header(sample))

    qr_x = x + width - padding - qr_size
    qr_y = y + 8 * mm
    c.drawImage(_build_qr_image(qr_data), qr_x, qr_y, width=qr_size, height=qr_size, preserveAspectRatio=True, mask='auto')

    text_x = x + padding
    max_text_width = width - qr_size - padding * 3
    label_cvc = sample.purchase_contract_cvc
    line_gap = 5.2 * mm
    text_lines = 4 + (1 if label_cvc else 0)
    text_block_height = text_lines * line_gap
    text_y = qr_y + (qr_size + text_block_height) / 2 - 3.5 * mm

    def draw_line(label, value, bold=False):
        nonlocal text_y
        c.setFillColor(HexColor("#111827"))
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 8.4 if bold else 7.4)
        text = f"{label}: {value or '-'}" if label else str(value or "-")
        if len(text) > 42:
            text = text[:39] + "..."
        c.drawString(text_x, text_y, text)
        text_y -= line_gap

    def draw_quality_line(value):
        nonlocal text_y
        eco, decaf = _quality_style_flags(value)
        text = f"Calidad: {value or '-'}"
        if len(text) > 42:
            text = text[:39] + "..."
        if eco or decaf:
            bg = "#F4DFC8" if decaf else "#DDEEDB"
            fg = "#5B351E" if decaf else "#244D2E"
            c.setFillColor(HexColor(bg))
            c.roundRect(text_x - 1.2 * mm, text_y - 2.4 * mm, max_text_width + 2.4 * mm, 4.6 * mm, 1.2 * mm, stroke=0, fill=1)
            c.setFillColor(HexColor(fg))
            c.setFont("Helvetica-Bold", 7.2)
            c.drawString(text_x, text_y, text)
            if eco and decaf:
                badge_w = 8 * mm
                badge_x = text_x + max_text_width - badge_w
                c.setFillColor(HexColor("#BFDDB8"))
                c.roundRect(badge_x, text_y - 2.1 * mm, badge_w, 4 * mm, 1 * mm, stroke=0, fill=1)
                c.setFillColor(HexColor("#244D2E"))
                c.setFont("Helvetica-Bold", 5.8)
                c.drawCentredString(badge_x + badge_w / 2, text_y - 0.45 * mm, "ECO")
        else:
            c.setFillColor(HexColor("#111827"))
            c.setFont("Helvetica", 7.4)
            c.drawString(text_x, text_y, text)
        text_y -= line_gap

    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(HexColor("#153F2B"))
    c.drawString(text_x, text_y, "Indian Ecotrade")
    text_y -= line_gap
    draw_quality_line(sample.quality)
    draw_line("Proveedor", sample.producer)
    draw_line("Ref. proveedor", sample.supplier_reference)
    if label_cvc:
        draw_line("CVC", label_cvc)
    c.restoreState()


def _render_avery_l7108rev_sheet(samples, request: Request, copies: int = 1, start_position: int = 1) -> io.BytesIO:
    layout = AVERY_L7108REV
    pdf_buffer = io.BytesIO()
    page_size = A4
    c = canvas.Canvas(pdf_buffer, pagesize=page_size)
    page_width, page_height = page_size
    base = os.getenv('APP_BASE_URL') or str(request.base_url)

    slots = []
    for row in range(layout["rows"]):
        for col in range(layout["columns"]):
            x = layout["margin_x"] + col * (layout["cell_width"] + layout["gap_x"])
            y = page_height - layout["margin_y"] - layout["cell_height"] - row * (layout["cell_height"] + layout["gap_y"])
            slots.append((x, y))

    current_slot = max(0, min(8, start_position - 1))
    for sample in samples:
        for _ in range(max(1, copies)):
            if current_slot >= len(slots):
                c.showPage()
                current_slot = 0
            x, y = slots[current_slot]
            qr_data = f"{base.rstrip('/')}/samples/{sample.id}"
            c.saveState()
            c.translate(x + layout["cell_width"], y)
            c.rotate(90)
            _draw_avery_l7108rev_label(c, sample, 0, 0, qr_data)
            c.restoreState()
            current_slot += 1

    c.save()
    pdf_buffer.seek(0)
    return pdf_buffer


@app.get("/samples/{sample_id}/label")
async def generate_label(sample_id: int, request: Request, db: Session = Depends(get_db)):
    """Generar hoja Avery L7108REV para una muestra."""
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")

    pdf_buffer = _render_avery_l7108rev_sheet([sample], request, copies=1, start_position=1)
    headers = {"Content-Disposition": f"attachment; filename=\"avery_l7108rev_{sample.code}.pdf\""}
    return StreamingResponse(pdf_buffer, media_type="application/pdf", headers=headers)


@app.get("/labels/pdf")
async def generate_labels_pdf(
    request: Request,
    ids: str,
    model: str = "L7108REV",
    copies: int = 1,
    start_position: int = 1,
    db: Session = Depends(get_db),
):
    """Generar hoja A4 Avery L7108REV con posicion inicial seleccionable."""
    if model != "L7108REV":
        raise HTTPException(status_code=400, detail="Modelo Avery no soportado")
    try:
        sample_ids = [int(value) for value in ids.split(",") if value.strip()]
    except ValueError:
        raise HTTPException(status_code=400, detail="IDs de muestra no validos")
    if not sample_ids:
        raise HTTPException(status_code=400, detail="Selecciona al menos una muestra")
    if copies < 1:
        raise HTTPException(status_code=400, detail="El numero de copias debe ser mayor que cero")
    if start_position < 1 or start_position > 9:
        raise HTTPException(status_code=400, detail="La posicion inicial debe estar entre 1 y 9")

    samples = db.query(Sample).filter(Sample.id.in_(sample_ids)).order_by(Sample.id).all()
    found_ids = {sample.id for sample in samples}
    missing = [sample_id for sample_id in sample_ids if sample_id not in found_ids]
    if missing:
        raise HTTPException(status_code=404, detail=f"Muestras no encontradas: {missing}")

    pdf_buffer = _render_avery_l7108rev_sheet(samples, request, copies=copies, start_position=start_position)
    headers = {"Content-Disposition": "attachment; filename=\"avery_l7108rev_etiquetas.pdf\""}
    return StreamingResponse(pdf_buffer, media_type="application/pdf", headers=headers)


@app.get("/samples/{sample_id}/label-old")
async def generate_label_old(sample_id: int, request: Request, db: Session = Depends(get_db)):
    """Generar etiqueta PDF con QR y banda de país"""
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")

    base = os.getenv('APP_BASE_URL')
    if not base:
        base = str(request.base_url)
    qr_data = f"{base.rstrip('/')}/samples/{sample_id}"
    qr = qrcode.QRCode(version=1, box_size=4, border=1)
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")

    pdf_buffer = io.BytesIO()
    page_width = 89 * mm
    page_height = 62 * mm
    c = canvas.Canvas(pdf_buffer, pagesize=(page_width, page_height))
    margin = 5 * mm

    band_height = 10 * mm
    country_colors = {
        'PE': HexColor('#DC143C'),
        'CO': HexColor('#FCD116'),
        'BR': HexColor('#009B3A'),
        'EC': HexColor('#F4D03F'),
        'BO': HexColor('#F4D03F'),
    }
    band_color = country_colors.get(sample.country_code, HexColor('#444444'))
    c.setFillColor(band_color)
    c.rect(0, page_height - band_height, page_width, band_height, fill=1, stroke=0)

    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 9)
    flag = get_country_flag(sample.country_code) if sample.country_code else '🌍'
    c.drawString(margin, page_height - band_height + 2 * mm, f"{flag} {sample.country_name or 'Unknown'}")

    qr_width = page_width * 0.38
    qr_height = qr_width
    info_width = page_width - qr_width - margin * 3
    info_x = margin
    qr_x = page_width - margin - qr_width
    y = page_height - band_height - margin

    c.setFillColor(HexColor('#000000'))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(info_x, y, "Indian Ecotrade")
    y -= 5 * mm
    c.setFont("Helvetica", 8)
    c.drawString(info_x, y, f"Origen: {sample.origin or '-'}")
    y -= 4.5 * mm
    c.drawString(info_x, y, f"Calidad: {sample.quality or '-'}")
    y -= 4.5 * mm
    c.drawString(info_x, y, f"Proveedor: {sample.producer or '-'}")
    y -= 4.5 * mm
    c.drawString(info_x, y, f"Ref. proveedor: {sample.supplier_reference or '-'}")
    y -= 4.5 * mm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(info_x, y, f"Código: {sample.code}")
    y -= 4.5 * mm
    if sample.purchase_contract_cvc:
        c.setFont("Helvetica", 8)
        c.drawString(info_x, y, f"CVC: {sample.purchase_contract_cvc}")
        y -= 4.5 * mm

    qr_path = io.BytesIO()
    qr_img.save(qr_path, format="PNG")
    qr_path.seek(0)
    try:
        c.drawImage(ImageReader(qr_path), qr_x, page_height - margin - qr_height, width=qr_width, height=qr_height, preserveAspectRatio=True, mask='auto')
    except Exception:
        pass

    c.save()
    pdf_buffer.seek(0)
    headers = {"Content-Disposition": f"attachment; filename=\"muestra_{sample.code}.pdf\""}
    return StreamingResponse(pdf_buffer, media_type="application/pdf", headers=headers)


# ============================================================================
# COMPARATOR
# ============================================================================

@app.get("/compare", response_class=HTMLResponse)
async def compare_samples(
    request: Request,
    ids: str = None,
    db: Session = Depends(get_db)
):
    """Comparador de muestras"""
    samples = []
    tastings = []
    
    if ids:
        sample_ids = [int(id) for id in ids.split(",")]
        samples = db.query(Sample).filter(Sample.id.in_(sample_ids)).all()
        tastings = {
            s.id: db.query(Tasting).filter(
                Tasting.sample_id == s.id
            ).order_by(desc(Tasting.indian_score)).first()
            for s in samples
        }
    
    return templates.TemplateResponse("compare.html", {
        "request": request,
        "samples": samples,
        "tastings": tastings,
        "all_samples": db.query(Sample).all(),
    })


@app.get('/labels', response_class=HTMLResponse)
async def labels_page(request: Request, ids: str = None, db: Session = Depends(get_db)):
    samples = db.query(Sample).order_by(desc(Sample.created_at)).all()
    selected_ids = set()
    if ids:
        for value in ids.split(","):
            value = value.strip()
            if value.isdigit():
                selected_ids.add(int(value))
    selected_samples = [sample for sample in samples if sample.id in selected_ids]
    return templates.TemplateResponse('labels.html', {
        "request": request,
        "all_samples": samples,
        "selected_label_ids": selected_ids,
        "selected_label_samples": selected_samples,
    })


@app.get('/tastings', response_class=HTMLResponse)
async def list_tastings(request: Request, db: Session = Depends(get_db)):
    tastings = db.query(Tasting).order_by(desc(Tasting.tasting_date)).all()
    return templates.TemplateResponse('tastings.html', {"request": request, "tastings": tastings})


@app.get('/samples/{sample_id}/tastings/new', response_class=HTMLResponse)
async def new_tasting_form(sample_id: int, request: Request, db: Session = Depends(get_db)):
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Muestra no encontrada")
    return templates.TemplateResponse('tasting_form.html', {"request": request, "sample": sample})


# ============================================================================
# IMPORT/EXPORT
# ============================================================================

IMPORT_HEADER_ALIASES = {
    "provider": {"provider", "proveedor", "supplier", "productor", "producer"},
    "supplier_reference": {"supplier_reference", "ref_proveedor", "referencia_proveedor", "ref proveedor", "referencia proveedor", "supplier ref", "supplier_ref"},
    "purchase_contract_cvc": {"purchase_contract_cvc", "cvc", "contrato_cvc", "contrato cvc", "purchase cvc"},
    "country": {"country", "pais", "país", "country_code", "origen pais"},
    "origin": {"origin", "origen", "region", "región", "finca"},
    "quality": {"quality", "calidad"},
    "quantity": {"quantity", "cantidad", "kg", "cantidad_kg"},
    "quantity_g": {"quantity_g", "cantidad_g", "gramos", "received_quantity_g", "cantidad recibida g"},
    "status": {"status", "estado"},
    "container_number": {"container_number", "contenedor", "numero_contenedor", "n contenedor", "nº contenedor", "no contenedor"},
    "warehouse": {"warehouse", "almacen", "almacén", "almacen warehouse", "almacén warehouse"},
    "tasting_date": {"fecha cata", "fehca cata", "tasting_date", "fecha_cata"},
    "code": {"code", "codigo", "código", "codigo interno", "internal code"},
    "variety": {"variety", "variedad"},
    "processing": {"processing", "proceso", "process"},
    "harvest_date": {"harvest_date", "cosecha", "harvest", "fecha cosecha"},
    "altitude": {"altitude", "altitud"},
    "notes": {"notes", "notas", "comments", "comentarios", "comentarios de los clientes", "comentarios clientes"},
}

IMPORT_SOURCE_SHEETS = {
    "Robusta America": "ROBUSTA_AMERICA",
    "Robusta Africa": "ROBUSTA_AFRICA",
    "Robusta Asia": "ROBUSTA_ASIA",
    "Descafeinados": "DESCAFEINADOS",
}


def _normalize_header(value):
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-/]+", " ", text)
    text = re.sub(r"[^a-z0-9áéíóúñüº ]", "", text)
    return text.strip()


def _canonical_header(value):
    normalized = _normalize_header(value)
    normalized_underscore = normalized.replace(" ", "_")
    for canonical, aliases in IMPORT_HEADER_ALIASES.items():
        if normalized in aliases or normalized_underscore in aliases:
            return canonical
    return normalized_underscore or None


def _clean_import_text(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"none", "null", "n/a", "na", "-", "--"}:
        return None
    return re.sub(r"\s+", " ", text)


def _clean_import_cvc(value):
    text = _clean_import_text(value)
    if not text:
        return None
    lowered = text.lower()
    invalid_tokens = ("oferta", "saco", "sacos", "bulto", "bultos", " tn", "ton", "tonelada", "cvv")
    if any(token in lowered for token in invalid_tokens):
        return None
    if "cvc" in lowered or lowered.startswith("c/c"):
        return text
    if re.search(r"\d{1,6}\s*[-/]\s*\d{2,4}", text):
        return text
    return None


def _clean_import_container(value):
    text = _clean_import_text(value)
    if not text:
        return None
    return re.sub(r"\s+", " ", text).upper()


def _identity_part(value):
    text = _clean_import_text(value)
    return text.upper() if text else None


def _make_import_identity(provider, supplier_reference, cvc, container_number=None):
    provider_key = _identity_part(provider)
    reference_key = _identity_part(supplier_reference)
    cvc_key = _identity_part(cvc)
    container_key = _identity_part(container_number)
    if not reference_key:
        return None, "incomplete"
    if cvc_key and container_key:
        return "|".join([provider_key or "", reference_key, cvc_key, container_key]), "primary_cvc_container"
    if cvc_key:
        return "|".join([provider_key or "", reference_key, cvc_key]), "primary_cvc"
    if container_key:
        return "|".join([provider_key or "", reference_key, container_key]), "secondary_reference_container"
    return "|".join([provider_key or "", reference_key]), "secondary_reference"


def _parse_import_quantity(row):
    def parse_number(value, multiplier=1):
        if value in (None, ""):
            return None
        text = str(value).strip().lower().replace(",", ".")
        match = re.search(r"(\d+(?:\.\d+)?)", text)
        if not match:
            return None
        return int(round(float(match.group(1)) * multiplier))

    quantity_g = row.get("quantity_g")
    if quantity_g not in (None, ""):
        return parse_number(quantity_g, 1)
    quantity = row.get("quantity")
    if quantity in (None, ""):
        return None
    text = str(quantity).strip().lower()
    multiplier = 1 if "gr" in text or "gram" in text else 1000
    return parse_number(quantity, multiplier)


def _normalize_import_status(value):
    if value in (None, ""):
        return SampleStatus.RECEIVED.value
    return normalize_status_value(str(value))


def _sample_identity(sample):
    return _make_import_identity(
        sample.producer,
        sample.supplier_reference,
        sample.purchase_contract_cvc,
        sample.container_number,
    )


def _similar_sample_exists(samples, normalized):
    provider = _identity_part(normalized.get("provider"))
    origin = _identity_part(normalized.get("origin"))
    quality = _identity_part(normalized.get("quality"))
    if not provider:
        return None
    for sample in samples:
        if _identity_part(sample.producer) != provider:
            continue
        same_origin = origin and _identity_part(sample.origin) == origin
        same_quality = quality and _identity_part(sample.quality) == quality
        if same_origin or same_quality:
            return sample
    return None


def _import_action_counts(rows):
    counts = {
        "CREATE_CANDIDATE": 0,
        "EXISTING_MATCH": 0,
        "DUPLICATE_IN_FILE": 0,
        "INCOMPLETE": 0,
        "WARNING_SIMILAR": 0,
        "ERROR": 0,
    }
    for row in rows:
        counts[row.proposed_action] = counts.get(row.proposed_action, 0) + 1
    return counts


def _generate_import_sample_code(db: Session, batch_id: int, row_id: int):
    base = f"IMP-{batch_id}-{row_id}"
    code = base[:50]
    suffix = 1
    while db.query(Sample).filter(Sample.code == code).first():
        tail = f"-{suffix}"
        code = f"{base[:50 - len(tail)]}{tail}"
        suffix += 1
    return code


def _skip_action_for_import_row(row: ImportRow):
    if row.proposed_action == "DUPLICATE_IN_FILE":
        return "SKIP_DUPLICATE"
    if row.proposed_action == "EXISTING_MATCH":
        return "SKIP_EXISTING"
    if row.proposed_action == "INCOMPLETE":
        return "SKIP_INCOMPLETE"
    if row.proposed_action in {"WARNING_SIMILAR", "ERROR"}:
        return "SKIP_ERROR"
    return "SKIP_MANUAL"


@app.get("/imports", response_class=HTMLResponse)
async def imports_index(request: Request, db: Session = Depends(get_db)):
    batches = db.query(ImportBatch).order_by(desc(ImportBatch.created_at)).limit(25).all()
    return templates.TemplateResponse("imports.html", {"request": request, "batches": batches})


@app.post("/imports/preview")
async def imports_preview(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"No se pudo cargar openpyxl: {exc}")

    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="El archivo esta vacio")

    batch = ImportBatch(
        filename=os.path.basename(file.filename or "import.xlsx"),
        file_hash=hashlib.sha256(contents).hexdigest(),
        status="preview",
    )
    db.add(batch)
    db.flush()

    import_dir = Path(UPLOADS_DIR) / "imports" / str(batch.id)
    import_dir.mkdir(parents=True, exist_ok=True)
    stored_path = import_dir / batch.filename
    stored_path.write_bytes(contents)
    batch.stored_file_path = str(stored_path)

    try:
        workbook = load_workbook(filename=io.BytesIO(contents), data_only=True)
    except Exception as exc:
        batch.status = "failed"
        db.commit()
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {exc}")

    selected_sheets = [
        (workbook[sheet_name], source_code)
        for sheet_name, source_code in IMPORT_SOURCE_SHEETS.items()
        if sheet_name in workbook.sheetnames
    ]
    if not selected_sheets:
        batch.status = "failed"
        db.commit()
        raise HTTPException(status_code=400, detail="No se encontraron pestañas Robusta America, Robusta Africa, Robusta Asia o Descafeinados")

    staged = []
    identity_counts = {}
    existing_samples = db.query(Sample).all()
    existing_by_identity = {}
    for sample in existing_samples:
        identity_key, identity_level = _sample_identity(sample)
        if identity_key and identity_level != "incomplete":
            existing_by_identity[identity_key] = sample

    for sheet, source_code in selected_sheets:
        header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), [])
        headers = [_canonical_header(value) for value in header_values]
        useful_headers = {header for header in headers if header}
        if not useful_headers:
            continue

        for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            raw = {}
            for index, value in enumerate(values):
                if index < len(headers) and headers[index]:
                    raw[headers[index]] = value
            if not any(value not in (None, "") for value in raw.values()):
                continue

            normalized = {
                "provider": _clean_import_text(raw.get("provider")),
                "supplier_reference": _clean_import_text(raw.get("supplier_reference")),
                "purchase_contract_cvc": _clean_import_cvc(raw.get("purchase_contract_cvc")),
                "country": _clean_import_text(raw.get("country")),
                "origin": _clean_import_text(raw.get("origin")),
                "quality": _clean_import_text(raw.get("quality")),
                "warehouse": _clean_import_text(raw.get("warehouse")),
                "tasting_date": _clean_import_text(raw.get("tasting_date")),
                "quantity_g": _parse_import_quantity(raw),
                "status": _normalize_import_status(raw.get("status")),
                "container_number": _clean_import_container(raw.get("container_number")),
                "code": _clean_import_text(raw.get("code")),
                "variety": _clean_import_text(raw.get("variety")),
                "processing": _clean_import_text(raw.get("processing")),
                "harvest_date": _clean_import_text(raw.get("harvest_date")),
                "altitude": _clean_import_text(raw.get("altitude")),
                "notes": _clean_import_text(raw.get("notes")),
            }
            row_errors = []
            has_quantity_value = raw.get("quantity") not in (None, "") or raw.get("quantity_g") not in (None, "")
            if has_quantity_value and normalized["quantity_g"] is None:
                row_errors.append("Cantidad no valida")

            identity_key, identity_level = _make_import_identity(
                normalized["provider"],
                normalized["supplier_reference"],
                normalized["purchase_contract_cvc"],
                normalized["container_number"],
            )
            if identity_key:
                identity_counts[identity_key] = identity_counts.get(identity_key, 0) + 1

            staged.append({
                "row_number": row_number,
                "source_sheet": source_code,
                "raw": raw,
                "normalized": normalized,
                "identity_key": identity_key,
                "identity_level": identity_level,
                "errors": row_errors,
            })

    import_rows = []
    for item in staged:
        errors = []
        warnings = []
        sample = None
        action = "CREATE_CANDIDATE"
        identity_key = item["identity_key"]
        normalized = item["normalized"]

        if item["errors"]:
            action = "ERROR"
            errors.extend(item["errors"])
        elif item["identity_level"] == "incomplete":
            action = "INCOMPLETE"
            errors.append("Falta referencia de proveedor; requiere revision manual")
        elif identity_key and identity_counts.get(identity_key, 0) > 1:
            action = "DUPLICATE_IN_FILE"
            errors.append("Identidad comercial duplicada dentro del Excel")
        elif identity_key in existing_by_identity:
            sample = existing_by_identity[identity_key]
            action = "EXISTING_MATCH"
            warnings.append(f"Coincide con muestra existente {sample.code}")
        else:
            similar = _similar_sample_exists(existing_samples, normalized)
            if similar:
                sample = similar
                action = "WARNING_SIMILAR"
                warnings.append(f"Posible similitud con {similar.code}; no se hara matching automatico")

        row_model = ImportRow(
            batch_id=batch.id,
            row_number=item["row_number"],
            source_sheet=item["source_sheet"],
            raw_data_json=json.dumps(item["raw"], ensure_ascii=False, default=str),
            normalized_data_json=json.dumps(normalized, ensure_ascii=False, default=str),
            identity_key=identity_key,
            identity_level=item["identity_level"],
            proposed_action=action,
            status="preview",
            sample_id=sample.id if sample and action == "EXISTING_MATCH" else None,
            errors_json=json.dumps(errors, ensure_ascii=False),
            warnings_json=json.dumps(warnings, ensure_ascii=False),
        )
        db.add(row_model)
        import_rows.append(row_model)

    counts = _import_action_counts(import_rows)
    batch.total_rows = len(import_rows)
    batch.create_count = counts.get("CREATE_CANDIDATE", 0)
    batch.existing_count = counts.get("EXISTING_MATCH", 0)
    batch.duplicate_count = counts.get("DUPLICATE_IN_FILE", 0)
    batch.incomplete_count = counts.get("INCOMPLETE", 0)
    batch.warning_count = counts.get("WARNING_SIMILAR", 0)
    batch.error_count = counts.get("ERROR", 0)
    db.commit()

    return RedirectResponse(f"/imports/{batch.id}", status_code=303)


@app.post("/imports/{batch_id}/apply")
async def imports_apply(
    batch_id: int,
    selected_row_ids: list[int] = Form(default=[]),
    row_ids: list[int] = Form(default=[]),
    candidate_row_ids: list[int] = Form(default=[]),
    db: Session = Depends(get_db),
):
    batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Importacion no encontrada")
    if batch.status not in {"preview", "partial_applied"}:
        raise HTTPException(status_code=400, detail="Esta importacion ya no esta en estado aplicable")

    selected_ids = set(selected_row_ids or row_ids or [])
    candidate_scope = set(candidate_row_ids or [])
    rows = db.query(ImportRow).filter(ImportRow.batch_id == batch_id).all()

    created_ids = []
    created_count = 0
    skipped_manual = 0
    skipped_duplicate = 0
    skipped_existing = 0
    skipped_incomplete = 0
    skipped_error = 0

    for row in rows:
        if row.status == "applied":
            if row.sample_id:
                created_ids.append(row.sample_id)
            continue

        if row.proposed_action == "CREATE_CANDIDATE":
            if row.id in selected_ids:
                data = json.loads(row.normalized_data_json or "{}")
                quantity_g = data.get("quantity_g") or 0
                status_value = data.get("status") or SampleStatus.RECEIVED.value
                sample = Sample(
                    code=_generate_import_sample_code(db, batch.id, row.id),
                    country_code=data.get("country"),
                    country_name=get_country_name(data.get("country")) if data.get("country") else None,
                    origin=data.get("origin") or "",
                    producer=data.get("provider") or "",
                    supplier_reference=data.get("supplier_reference"),
                    container_number=data.get("container_number"),
                    purchase_contract_cvc=data.get("purchase_contract_cvc"),
                    quality=data.get("quality"),
                    warehouse=data.get("warehouse"),
                    harvest_date=data.get("harvest_date") or "",
                    variety=data.get("variety") or "",
                    altitude=int(data.get("altitude")) if str(data.get("altitude") or "").isdigit() else 0,
                    processing=data.get("processing") or "",
                    initial_quantity=(quantity_g or 0) / 1000,
                    available_quantity=(quantity_g or 0) / 1000,
                    received_quantity_g=quantity_g or 0,
                    available_quantity_g=quantity_g or 0,
                    notes=data.get("notes"),
                    status=status_value if status_value in {s.value for s in SampleStatus} else SampleStatus.RECEIVED,
                )
                db.add(sample)
                db.flush()
                row.sample_id = sample.id
                row.final_action = "APPLY_CREATE"
                row.status = "applied"
                created_ids.append(sample.id)
                created_count += 1

                event = Event(
                    sample_id=sample.id,
                    event_type="received",
                    description=f"Muestra creada desde importacion #{batch.id} fila {row.row_number}",
                )
                db.add(event)
            elif row.id in candidate_scope:
                row.final_action = "SKIP_MANUAL"
                row.status = "skipped"
                skipped_manual += 1
            continue

        if not row.final_action:
            row.final_action = _skip_action_for_import_row(row)
            row.status = "skipped"
            if row.final_action == "SKIP_DUPLICATE":
                skipped_duplicate += 1
            elif row.final_action == "SKIP_EXISTING":
                skipped_existing += 1
            elif row.final_action == "SKIP_INCOMPLETE":
                skipped_incomplete += 1
            else:
                skipped_error += 1

    db.flush()
    pending = db.query(ImportRow).filter(
        ImportRow.batch_id == batch_id,
        ImportRow.proposed_action == "CREATE_CANDIDATE",
        ImportRow.final_action.is_(None),
    ).count()
    batch.status = "applied" if pending == 0 else "partial_applied"
    if batch.status == "applied":
        batch.applied_at = datetime.utcnow()

    db.commit()

    params = f"?created={created_count}&manual={skipped_manual}&duplicate={skipped_duplicate}&existing={skipped_existing}&incomplete={skipped_incomplete}&errors={skipped_error}"
    return RedirectResponse(f"/imports/{batch_id}{params}", status_code=303)


@app.post("/imports/{batch_id}/delete-created")
async def imports_delete_created(batch_id: int, db: Session = Depends(get_db)):
    batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Importacion no encontrada")

    rows = db.query(ImportRow).filter(
        ImportRow.batch_id == batch_id,
        ImportRow.final_action == "APPLY_CREATE",
        ImportRow.sample_id.isnot(None),
    ).all()
    sample_ids = sorted({row.sample_id for row in rows if row.sample_id})
    if not sample_ids:
        return RedirectResponse(f"/imports/{batch_id}?deleted=0", status_code=303)

    samples = db.query(Sample).filter(Sample.id.in_(sample_ids)).all()
    deleted_count = 0
    for sample in samples:
        _delete_sample_records(db, sample, import_final_action="DELETED_BY_BATCH_ROLLBACK", import_batch_id=batch.id)
        deleted_count += 1

    batch.status = "rolled_back_simple"
    batch.rolled_back_at = datetime.utcnow()
    db.commit()
    return RedirectResponse(f"/imports/{batch_id}?deleted={deleted_count}", status_code=303)


@app.get("/imports/{batch_id}", response_class=HTMLResponse)
async def imports_preview_detail(
    batch_id: int,
    request: Request,
    action: str = None,
    created: int = None,
    manual: int = None,
    duplicate: int = None,
    existing: int = None,
    incomplete: int = None,
    errors: int = None,
    deleted: int = None,
    db: Session = Depends(get_db),
):
    batch = db.query(ImportBatch).filter(ImportBatch.id == batch_id).first()
    if not batch:
        raise HTTPException(status_code=404, detail="Importacion no encontrada")
    rows_query = db.query(ImportRow).filter(ImportRow.batch_id == batch_id)
    if action:
        rows_query = rows_query.filter(ImportRow.proposed_action == action)
    rows = rows_query.order_by(ImportRow.source_sheet, ImportRow.row_number).all()
    parsed_rows = []
    for row in rows:
        parsed_rows.append({
            "row": row,
            "normalized": json.loads(row.normalized_data_json or "{}"),
            "errors": json.loads(row.errors_json or "[]"),
            "warnings": json.loads(row.warnings_json or "[]"),
        })
    duplicate_rows = db.query(ImportRow).filter(
        ImportRow.batch_id == batch_id,
        ImportRow.proposed_action == "DUPLICATE_IN_FILE",
    ).order_by(ImportRow.identity_key, ImportRow.source_sheet, ImportRow.row_number).all()
    duplicate_groups_by_key = {}
    for row in duplicate_rows:
        data = json.loads(row.normalized_data_json or "{}")
        group = duplicate_groups_by_key.setdefault(row.identity_key or "", {
            "identity_key": row.identity_key or "-",
            "count": 0,
            "sheets": set(),
            "rows": [],
            "provider": data.get("provider"),
            "supplier_reference": data.get("supplier_reference"),
            "purchase_contract_cvc": data.get("purchase_contract_cvc"),
            "quality": data.get("quality"),
            "origin": data.get("origin"),
        })
        group["count"] += 1
        group["sheets"].add(row.source_sheet or "-")
        group["rows"].append(f"{row.source_sheet or '-'}:{row.row_number}")
    duplicate_groups = []
    for group in duplicate_groups_by_key.values():
        group["sheets"] = sorted(group["sheets"])
        duplicate_groups.append(group)

    created_sample_ids = [
        str(sample_id)
        for (sample_id,) in db.query(ImportRow.sample_id)
        .filter(
            ImportRow.batch_id == batch_id,
            ImportRow.final_action == "APPLY_CREATE",
            ImportRow.sample_id.isnot(None),
        )
        .order_by(ImportRow.id)
        .all()
    ]
    pending_create_count = db.query(ImportRow).filter(
        ImportRow.batch_id == batch_id,
        ImportRow.proposed_action == "CREATE_CANDIDATE",
        ImportRow.final_action.is_(None),
    ).count()
    applied_create_count = db.query(ImportRow).filter(
        ImportRow.batch_id == batch_id,
        ImportRow.final_action == "APPLY_CREATE",
    ).count()
    manual_skip_count = db.query(ImportRow).filter(
        ImportRow.batch_id == batch_id,
        ImportRow.final_action == "SKIP_MANUAL",
    ).count()
    apply_summary = None
    if any(value is not None for value in [created, manual, duplicate, existing, incomplete, errors]):
        apply_summary = {
            "created": created or 0,
            "manual": manual or 0,
            "duplicate": duplicate or 0,
            "existing": existing or 0,
            "incomplete": incomplete or 0,
            "errors": errors or 0,
        }
    rollback_summary = None
    if deleted is not None:
        rollback_summary = {"deleted": deleted or 0}
    return templates.TemplateResponse("imports_preview.html", {
        "request": request,
        "batch": batch,
        "rows": parsed_rows,
        "active_action": action or "",
        "duplicate_groups": duplicate_groups,
        "apply_summary": apply_summary,
        "created_sample_ids": created_sample_ids,
        "created_sample_ids_csv": ",".join(created_sample_ids),
        "pending_create_count": pending_create_count,
        "applied_create_count": applied_create_count,
        "manual_skip_count": manual_skip_count,
        "rollback_summary": rollback_summary,
    })


@app.post("/import/excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Importar muestras desde Excel"""
    try:
        from openpyxl import load_workbook
        
        contents = await file.read()
        wb = load_workbook(filename=io.BytesIO(contents))
        ws = wb.active
        
        imported = 0
        errors = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # Try to unpack common columns; handle extra optional columns
                values = list(row)
                code = values[0] if len(values) > 0 else None
                country = values[1] if len(values) > 1 else None
                origin = values[2] if len(values) > 2 else None
                producer = values[3] if len(values) > 3 else None
                harvest = values[4] if len(values) > 4 else None
                variety = values[5] if len(values) > 5 else None
                altitude = values[6] if len(values) > 6 else None
                processing = values[7] if len(values) > 7 else None
                quantity = values[8] if len(values) > 8 else None
                # extra fields
                supplier_reference = values[9] if len(values) > 9 else None
                provider_sample_number = values[10] if len(values) > 10 else None
                purchase_cvc = values[11] if len(values) > 11 else None
                sales_cvv = values[12] if len(values) > 12 else None
                quality = values[13] if len(values) > 13 else None
                warehouse = values[14] if len(values) > 14 else None
                sample_type = values[15] if len(values) > 15 else None
                category = values[16] if len(values) > 16 else None
                comments = values[17] if len(values) > 17 else None

                if not code:
                    continue

                existing = db.query(Sample).filter(Sample.code == code).first()
                if existing:
                    errors.append(f"Row {row_idx}: Código duplicado")
                    continue

                sample = Sample(
                    code=str(code).strip(),
                    country_code=str(country).strip()[:5] if country else None,
                    country_name=get_country_name(str(country).strip()[:5]) if country else None,
                    origin=str(origin).strip() if origin else "",
                    producer=str(producer).strip() if producer else "",
                    supplier_reference=str(supplier_reference).strip() if supplier_reference else None,
                    provider_sample_number=str(provider_sample_number).strip() if provider_sample_number else None,
                    purchase_contract_cvc=str(purchase_cvc).strip() if purchase_cvc else None,
                    sales_contract_cvv=str(sales_cvv).strip() if sales_cvv else None,
                    quality=str(quality).strip() if quality else None,
                    warehouse=str(warehouse).strip() if warehouse else None,
                    sample_type=str(sample_type).strip() if sample_type else None,
                    category=str(category).strip() if category else None,
                    notes=str(comments).strip() if comments else None,
                    harvest_date=str(harvest).strip() if harvest else "",
                    variety=str(variety).strip() if variety else "",
                    altitude=int(altitude) if altitude else 0,
                    processing=str(processing).strip() if processing else "",
                    initial_quantity=float(quantity) if quantity else 0,
                    available_quantity=float(quantity) if quantity else 0,
                )

                db.add(sample)
                imported += 1

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")
        
        db.commit()
        
        return JSONResponse({
            "imported": imported,
            "errors": errors
        })
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/import", response_class=HTMLResponse)
async def import_form(request: Request):
    return templates.TemplateResponse("import.html", {"request": request})


# ============================================================================
# API ENDPOINTS
# ============================================================================

@app.get("/api/samples")
async def api_samples(db: Session = Depends(get_db)):
    """API: Lista de muestras"""
    samples = db.query(Sample).all()
    return {
        "samples": [
            {
                "id": s.id,
                "code": s.code,
                "country": s.country_name,
                "origin": s.origin,
                "available": s.available_quantity,
                "status": s.status.value,
            }
            for s in samples
        ]
    }


@app.get("/api/samples/{sample_id}")
async def api_sample_detail(sample_id: int, db: Session = Depends(get_db)):
    """API: Detalle de muestra"""
    sample = db.query(Sample).filter(Sample.id == sample_id).first()
    if not sample:
        raise HTTPException(status_code=404, detail="Not found")
    
    tastings = db.query(Tasting).filter(Tasting.sample_id == sample_id).all()
    
    return {
        "id": sample.id,
        "code": sample.code,
        "country": sample.country_name,
        "origin": sample.origin,
        "producer": sample.producer,
        "available": sample.available_quantity,
        "status": sample.status.value,
        "tastings": [
            {
                "id": t.id,
                "evaluator": t.evaluator,
                "date": t.tasting_date.isoformat(),
                "indian_score": t.indian_score,
                "cup_score": t.cup_score,
                "commercial_score": t.commercial_score,
            }
            for t in tastings
        ]
    }


@app.get("/api/health")
async def health():
    """Health check"""
    return {"status": "ok"}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
